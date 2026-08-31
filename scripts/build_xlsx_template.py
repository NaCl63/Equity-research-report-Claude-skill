#!/usr/bin/env python3
"""
Equity research model template (openpyxl).

HOW TO USE THIS FILE:
  1. Copy it into your working directory and rename it for the ticker you're building
     (e.g. build_xlsx_acme.py).
  2. Search for "TICKER-SPECIFIC" — every block marked that way holds example/placeholder
     data that must be replaced with real, sourced numbers for your target company.
  3. Keep the helper functions (style constants, `cell()`, `section_bar()`, `data_table`-style
     row builders, named-range setup) as-is — they encode formatting conventions and the
     gotchas documented in SKILL.md (never start a plain-text cell with "=", track row numbers
     via variables not hand counts, wrap margins in ISBLANK-aware IFERROR, etc).
  4. After building, run:
       recalc the workbook to zero formula errors, then
       python3 scripts/verify_model.py <output.xlsx> --price <current price>
     and do not deliver the file until BOTH are clean. verify_model.py is the second half of
     that gate: a clean recalc proves formulas *evaluate*, not that they reference the right
     cells or produce sane numbers.

Tabs built here: Cover, Quarterly Financials, Driver Sensitivity, Valuation, Analyst Ratings,
Technical & Market, Catalyst Calendar, Sources & Notes. Add/remove tabs to fit the company —
this structure fit a stablecoin issuer (Circle/CRCL) well; a bank, a SaaS company, or a miner
will want a different driver-sensitivity tab in particular.
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule

# ============================================================================
# STYLE CONSTANTS — keep these conventions; they match the xlsx skill's
# "Financial models" section (blue=input, black=formula, green=cross-sheet link).
# ============================================================================
FONT_NAME = "Arial"

BLUE = Font(name=FONT_NAME, color="0000FF", size=10)          # hardcoded inputs
BLACK = Font(name=FONT_NAME, color="000000", size=10)          # formulas
GREEN = Font(name=FONT_NAME, color="008000", size=10)          # links to other sheet
BOLD = Font(name=FONT_NAME, bold=True, size=10)
BOLD_BLUE = Font(name=FONT_NAME, bold=True, color="0000FF", size=10)
TITLE = Font(name=FONT_NAME, bold=True, size=16, color="1F3864")
SUBTITLE = Font(name=FONT_NAME, size=11, italic=True, color="595959")
HDR = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
SECTION = Font(name=FONT_NAME, bold=True, size=12, color="FFFFFF")
NOTE = Font(name=FONT_NAME, italic=True, size=9, color="595959")

HDR_FILL = PatternFill("solid", fgColor="1F3864")
SECTION_FILL = PatternFill("solid", fgColor="2E5395")
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")   # key assumptions / levers to change
LIGHT_FILL = PatternFill("solid", fgColor="D9E2F3")
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CUR0 = '$#,##0;($#,##0);"-"'
CUR1 = '$#,##0.0;($#,##0.0);"-"'
CUR2 = '$#,##0.00;($#,##0.00);"-"'
PCT1 = '0.0%;(0.0%);"-"'
PCT0 = '0%;(0%);"-"'
MULT = '0.00"x"'
NUM0 = '#,##0;(#,##0);"-"'


def section_bar(ws, row, text, span, start_col=1):
    """Full-width navy section header bar."""
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + span - 1)
    c = ws.cell(row=row, column=start_col, value=text)
    c.font = SECTION
    c.fill = SECTION_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def cell(ws, row, col, value, font=BLACK, fmt=None, fill=None, align=None, border=False, wrap=False):
    """Write one cell with consistent styling. Returns the cell object."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if align or wrap:
        c.alignment = Alignment(horizontal=align or "general", wrap_text=wrap, vertical="center")
    if border:
        c.border = BORDER
    return c


def header_row(ws, row, headers, start_col=2):
    """Navy header row across the given columns."""
    for i, h in enumerate(headers, start=start_col):
        c = cell(ws, row, i, h, HDR, fill=HDR_FILL, align="center", border=True, wrap=True)
    ws.row_dimensions[row].height = 18


wb = Workbook()

# ============================================================================
# TAB 1 — COVER
#   TICKER-SPECIFIC: replace TICKER, COMPANY_NAME, and every snapshot stat below.
#   Keep the named-range pattern (Cover_Price / Cover_Shares / Cover_MktCap) — other
#   tabs should reference these by name, not by a hardcoded cell coordinate.
# ============================================================================
TICKER = "TICK"
COMPANY_NAME = "EXAMPLE COMPANY, INC."

ws = wb.active
ws.title = "Cover"
set_col_widths(ws, [3, 26, 20, 20, 20, 20, 3])
ws.sheet_view.showGridLines = False

cell(ws, 2, 2, COMPANY_NAME, TITLE)
cell(ws, 3, 2, f"{TICKER} — Equity Research Model — Fundamental + Technical", SUBTITLE)
cell(ws, 4, 2, "Prepared for informational/educational purposes — not investment advice", NOTE)
cell(ws, 5, 2, "Model date:", BOLD)
cell(ws, 5, 3, "YYYY-MM-DD", BLUE)  # TICKER-SPECIFIC: report date

section_bar(ws, 7, "MARKET SNAPSHOT", 5)
# TICKER-SPECIFIC: every value below — source each one and note the source in column E.
snap = [
    ("Last price", 0.00, CUR2, "source"),
    ("Prior close", 0.00, CUR2, "source"),
    ("52-week high", 0.00, CUR2, "source"),
    ("52-week low", 0.00, CUR2, "source"),
    ("Shares outstanding (mm)", 0.0, NUM0, "source"),
    ("Float (mm)", 0.0, NUM0, "source"),
    ("Market capitalization ($mm)", None, CUR0, "Formula: price x shares out"),
    ("Enterprise value ($mm)", 0, CUR0, "source"),
]
price_row = 8
r = price_row
for i, (label, val, fmt, src) in enumerate(snap):
    this_row = price_row + i
    cell(ws, this_row, 2, label, BLACK)
    if label.startswith("Market cap"):
        cell(ws, this_row, 3, "=Cover_Price*Cover_Shares", BLACK, fmt)
    else:
        cell(ws, this_row, 3, val, BLUE, fmt)
    cell(ws, this_row, 5, src, NOTE)
    if label.startswith("Shares outstanding"):
        shares_row = this_row
    if label.startswith("Market cap"):
        mcap_row = this_row
    r = this_row

wb.defined_names["Cover_Price"] = DefinedName("Cover_Price", attr_text=f"Cover!$C${price_row}")
wb.defined_names["Cover_Shares"] = DefinedName("Cover_Shares", attr_text=f"Cover!$C${shares_row}")
wb.defined_names["Cover_MktCap"] = DefinedName("Cover_MktCap", attr_text=f"Cover!$C${mcap_row}")

section_bar(ws, r + 2, "REPORT CONTENTS", 5)
contents = [
    "1. Quarterly Financials — historical trend + margin build",
    "2. Driver Sensitivity — the single biggest valuation lever, modeled explicitly",
    "3. Valuation — comps table + scenario (bear/base/bull) price targets",
    "4. Analyst Ratings — current sell-side coverage",
    "5. Technical & Market Data — price levels, options, short interest, ownership",
    "6. Catalyst Calendar — historical and forward catalysts",
    "7. Sources & Notes — citations and data-quality caveats",
]
rr = r + 3
for line in contents:
    cell(ws, rr, 2, line, BLACK)
    rr += 1

cell(ws, rr + 1, 2, "Color key:", BOLD)
cell(ws, rr + 2, 2, "Blue = hardcoded input/assumption", BLUE)
cell(ws, rr + 3, 2, "Black = formula", BLACK)
cell(ws, rr + 4, 2, "Green = link to another sheet", GREEN)
cell(ws, rr + 5, 2, "Yellow fill = key lever — change this to re-run scenarios", BLACK, fill=YELLOW_FILL)

cell(ws, rr + 7, 2,
     "DISCLAIMER: This workbook was built from public secondary sources as of the model date "
     "above, not a live market data terminal. Every hardcoded figure should be sourced on the "
     "Sources & Notes tab. Flag any figure that showed variance across sources inline rather "
     "than silently picking one. Nothing here is investment, legal, or tax advice.",
     NOTE, wrap=True)
ws.row_dimensions[rr + 7].height = 55
ws.merge_cells(start_row=rr + 7, start_column=2, end_row=rr + 7, end_column=6)

# ============================================================================
# TAB 2 — QUARTERLY FINANCIALS
#   TICKER-SPECIFIC: replace the metric rows and hardcoded quarters with the
#   company's actual reported line items. Keep at least 3-5 recent quarters so a
#   deceleration/acceleration trend is visible, not just annual totals.
# ============================================================================
ws2 = wb.create_sheet("Quarterly Financials")
set_col_widths(ws2, [2, 34, 15, 15, 15, 15, 15, 3])
ws2.sheet_view.showGridLines = False

section_bar(ws2, 2, f"{TICKER} — HISTORICAL FINANCIALS ($mm unless noted)", 6)
cell(ws2, 3, 2, "Source: [cite company press releases / 10-Q / 10-K here]. Note any derived "
     "(not directly disclosed) rows explicitly.", NOTE, wrap=True)
ws2.merge_cells(start_row=3, start_column=2, end_row=3, end_column=7)
ws2.row_dimensions[3].height = 28

hdr_row = 5
period_labels = ["FYPrior A", "FYCurrent A", "QA", "QB", "QC"]  # TICKER-SPECIFIC: real period labels
header_row(ws2, hdr_row, ["Line item"] + period_labels)

r = hdr_row + 1


def fin_row(label, vals, fmt=CUR0, bold=False, is_formula=False, note=None):
    """Write one financial-statement row across the period columns C:G. Returns the row #."""
    global r
    this_row = r
    f = BOLD if bold else (BLACK if is_formula else BLUE)
    cell(ws2, this_row, 2, label, BOLD if bold else BLACK)
    for i, v in enumerate(vals or []):
        cell(ws2, this_row, 3 + i, v, f, fmt)
    if note:
        cell(ws2, this_row, 8, note, NOTE)
    r += 1
    return this_row


# TICKER-SPECIFIC block — replace every metric and value.
R_totrev = fin_row("Total revenue", [0, 0, 0, 0, 0], CUR0)
R_costs = fin_row("Cost of revenue / key cost line", [0, 0, 0, 0, 0], CUR0)
R_grossprofit = fin_row("Gross profit (= revenue - costs)", None, CUR0, is_formula=True)
for i, col in enumerate(["C", "D", "E", "F", "G"]):
    ws2.cell(row=R_grossprofit, column=3 + i).value = f"={col}{R_totrev}-{col}{R_costs}"
    ws2.cell(row=R_grossprofit, column=3 + i).font = BLACK
    ws2.cell(row=R_grossprofit, column=3 + i).number_format = CUR0

R_grossmargin = fin_row("  Gross margin", None, PCT1, is_formula=True)
for i, col in enumerate(["C", "D", "E", "F", "G"]):
    ws2.cell(row=R_grossmargin, column=3 + i).value = (
        f"=IF(ISBLANK({col}{R_totrev}),\"n/a\",IFERROR({col}{R_grossprofit}/{col}{R_totrev},\"-\"))"
    )
    ws2.cell(row=R_grossmargin, column=3 + i).font = BLACK

R_ni = fin_row("Net income / (loss)", [0, 0, 0, 0, 0], CUR0, note="Flag any one-time items here")
R_ebitda = fin_row("Adjusted EBITDA", [None, 0, 0, 0, 0], CUR0)
R_ebitdamargin = fin_row("  Adj. EBITDA margin", None, PCT1, is_formula=True)
for i, col in enumerate(["C", "D", "E", "F", "G"]):
    ws2.cell(row=R_ebitdamargin, column=3 + i).value = (
        f"=IF(ISBLANK({col}{R_ebitda}),\"n/a\",IFERROR({col}{R_ebitda}/{col}{R_totrev},\"-\"))"
    )
    ws2.cell(row=R_ebitdamargin, column=3 + i).font = BLACK

r += 1
R_kpi = fin_row("Key operating KPI (e.g. users, GMV, circulation)", [0, 0, 0, 0, 0], NUM0)
R_revyoy = fin_row("Revenue YoY growth (as disclosed)", [None, 0, 0, 0, 0], PCT0)

for rr_ in range(hdr_row + 1, r):
    for cc in range(2, 8):
        ws2.cell(row=rr_, column=cc).border = BORDER
        if (rr_ - hdr_row) % 2 == 0 and ws2.cell(row=rr_, column=cc).fill.fgColor.rgb in (None, "00000000"):
            ws2.cell(row=rr_, column=cc).fill = ALT_FILL

cell(ws2, r + 1, 2,
     "Read-through: [state the deceleration/acceleration/margin story the numbers actually "
     "tell — this sentence is the whole point of building the table].", NOTE, wrap=True)
ws2.merge_cells(start_row=r + 1, start_column=2, end_row=r + 1, end_column=7)
ws2.row_dimensions[r + 1].height = 42

# ---- Native Excel charts on the trend data -------------------------------------------------
# Deliberately NATIVE openpyxl charts, not an image pasted in: they reference the cells, so when
# someone changes a figure or extends the period columns the chart follows. A picture of a chart
# in a live model goes stale the first time anyone touches an input.
#
# Two charts because they answer different questions and share no sensible y-axis: dollars of
# revenue and EBITDA on one, percentage growth and margin on the other. Forcing percentages onto
# a dollar axis is the most common way a financial chart ends up unreadable.
_n_periods = len(period_labels)
_cats = Reference(ws2, min_col=3, max_col=2 + _n_periods, min_row=hdr_row)

_ch1 = BarChart()
_ch1.type, _ch1.grouping = "col", "clustered"
_ch1.title = "Revenue and adjusted EBITDA"
_ch1.y_axis.title = "$mm"
_ch1.height, _ch1.width = 7.5, 16
_ch1.add_data(Reference(ws2, min_col=2, max_col=2 + _n_periods, min_row=R_totrev, max_row=R_totrev),
              titles_from_data=True, from_rows=True)
_ch1.add_data(Reference(ws2, min_col=2, max_col=2 + _n_periods, min_row=R_ebitda, max_row=R_ebitda),
              titles_from_data=True, from_rows=True)
_ch1.set_categories(_cats)
ws2.add_chart(_ch1, f"B{r + 4}")

_ch2 = LineChart()
_ch2.title = "Gross margin and adjusted EBITDA margin"
_ch2.y_axis.title = "% of revenue"
_ch2.y_axis.numFmt = "0.0%"
_ch2.height, _ch2.width = 7.5, 16
_ch2.add_data(Reference(ws2, min_col=2, max_col=2 + _n_periods, min_row=R_grossmargin, max_row=R_grossmargin),
              titles_from_data=True, from_rows=True)
_ch2.add_data(Reference(ws2, min_col=2, max_col=2 + _n_periods, min_row=R_ebitdamargin, max_row=R_ebitdamargin),
              titles_from_data=True, from_rows=True)
_ch2.set_categories(_cats)
for _s in _ch2.series:
    _s.smooth = False   # smoothed lines invent inflection points between real data points
ws2.add_chart(_ch2, f"B{r + 20}")

ws2.freeze_panes = f"C{hdr_row + 1}"

# ============================================================================
# TAB 3 — DRIVER SENSITIVITY
#   TICKER-SPECIFIC: this whole tab. Identify the ONE variable that drives valuation
#   (rate spread, commodity price, churn, take-rate, whatever) and build a small
#   2-D sensitivity grid around it, exactly like this skeleton does. This is the
#   single highest-value tab in the whole model — don't skip it or leave it generic.
# ============================================================================
ws3 = wb.create_sheet("Driver Sensitivity")
set_col_widths(ws3, [2, 30, 13, 13, 13, 13, 13, 13, 13, 3])
ws3.sheet_view.showGridLines = False

section_bar(ws3, 2, "KEY DRIVER MODEL & SENSITIVITY", 8)
cell(ws3, 3, 2,
     "[Name the driver and explain in one sentence why it dominates the valuation. Example: "
     "\"Circle's revenue is ~95%+ reserve/interest income on USDC-backing Treasuries.\"] "
     "Change the yellow cells to re-run the grid below.", NOTE, wrap=True)
ws3.merge_cells(start_row=3, start_column=2, end_row=3, end_column=9)
ws3.row_dimensions[3].height = 28

section_bar(ws3, 5, "BASE CASE ASSUMPTIONS", 8)
assum_row = 6
cell(ws3, assum_row, 2, "Base value of driver input #1 (e.g. volume/circulation/units)", BLACK)
cell(ws3, assum_row, 4, 0, BLUE, CUR1, fill=YELLOW_FILL)
cell(ws3, assum_row, 6, "source", NOTE)
BASE_A = f"D{assum_row}"

assum_row += 1
cell(ws3, assum_row, 2, "Base value of driver input #2 (e.g. rate/spread/margin)", BLACK)
cell(ws3, assum_row, 4, 0, BLUE, PCT1, fill=YELLOW_FILL)
cell(ws3, assum_row, 6, "source or derivation", NOTE)
BASE_B = f"D{assum_row}"

section_bar(ws3, assum_row + 2, "SENSITIVITY GRID — vary input #2 (rows) x input #1 growth (columns)", 8)
grid_hdr = assum_row + 3
cell(ws3, grid_hdr, 2, "Input #2 change ↓  /  Input #1 growth →", BOLD, wrap=True)
growth_scenarios = [-0.10, 0.0, 0.10, 0.20, 0.30, 0.40]  # TICKER-SPECIFIC: pick sensible range
for i, g in enumerate(growth_scenarios):
    cell(ws3, grid_hdr, 3 + i, g, BOLD_BLUE, PCT0, fill=LIGHT_FILL, align="center")

rate_scenarios = [-100, -75, -50, -25, 0, 25]  # TICKER-SPECIFIC: e.g. bps, or $/unit, or pp of churn
grid_start = grid_hdr + 1
for j, chg in enumerate(rate_scenarios):
    rr_ = grid_start + j
    # The row axis is written as a NUMBER and referenced by the formulas ($B{row}), not baked into
    # them as a Python literal. Writing the value twice — once as a label, once inside the formula
    # string — means editing the visible label changes nothing, which is a silent, confidence-
    # destroying failure in a grid whose whole job is to be adjusted by the reader.
    cell(ws3, rr_, 2, chg, BOLD_BLUE, '+#,##0;-#,##0;0', fill=LIGHT_FILL, align="center")
    for i, g in enumerate(growth_scenarios):
        col = get_column_letter(3 + i)
        f = f"=({BASE_B}+$B{rr_}/10000)*({BASE_A}*(1+{col}${grid_hdr}))*1000"
        cell(ws3, rr_, 3 + i, f, BLACK, CUR0, border=True)

# Heat map: the shape of a sensitivity grid is the finding. A wall of same-looking numbers hides
# whether the model is more sensitive to the row driver or the column driver; a color scale makes
# the answer visible in the direction of the gradient.
ws3.conditional_formatting.add(
    f"C{grid_start}:{get_column_letter(2 + len(growth_scenarios))}{grid_start + len(rate_scenarios) - 1}",
    ColorScaleRule(start_type="min", start_color="F8696B",
                   mid_type="percentile", mid_value=50, mid_color="FFEB84",
                   end_type="max", end_color="63BE7B"))

note_row = grid_start + len(rate_scenarios) + 1
cell(ws3, note_row, 2,
     "[Cross-check this grid's base-case cell against any independent estimate you found in "
     "research — e.g. an analyst's or investor's back-of-envelope sensitivity claim — and note "
     "agreement or disagreement here.]", NOTE, wrap=True)
ws3.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=9)
ws3.row_dimensions[note_row].height = 42

cell(ws3, note_row + 2, 2,
     "Model limitation: [state the simplifying assumptions explicitly — e.g. assumes 1:1 "
     "pass-through, constant cost ratio, no competitive response — so the reader knows this is "
     "a directional framework, not a precision forecast].", NOTE, wrap=True)
ws3.merge_cells(start_row=note_row + 2, start_column=2, end_row=note_row + 2, end_column=9)
ws3.row_dimensions[note_row + 2].height = 42

# ============================================================================
# TAB 4 — VALUATION
#   TICKER-SPECIFIC: real comps and real scenario assumptions.
# ============================================================================
ws4 = wb.create_sheet("Valuation")
set_col_widths(ws4, [2, 26, 13, 13, 13, 13, 13, 13, 3])
ws4.sheet_view.showGridLines = False

section_bar(ws4, 2, "COMPARABLE COMPANY MULTIPLES", 7)
hdr = 4
header_row(ws4, hdr, ["Company", "Price", "Mkt Cap ($bn)", "EV ($bn)", "TTM Rev ($bn)", "EV/Rev (TTM)", "Fwd P/E"])

# comps: (name, price, mktcap_or_None, ev_or_None, ttm_rev, fwd_pe, is_subject_company)
comps = [
    (f"{COMPANY_NAME} ({TICKER})", 0.00, None, 0.0, 0.0, 0.0, True),
    ("Comp A", 0.00, 0.0, 0.0, 0.0, 0.0, False),
    ("Comp B", 0.00, 0.0, 0.0, 0.0, 0.0, False),
]
rr = hdr + 1
for name, price, mcap, ev, rev, pe, is_subject in comps:
    cell(ws4, rr, 2, name, BOLD if is_subject else BLACK, fill=LIGHT_FILL if is_subject else None)
    cell(ws4, rr, 3, price, BLUE, CUR2, border=True)
    if is_subject:
        cell(ws4, rr, 4, "=Cover_MktCap/1000", BLACK, CUR1, border=True)
    else:
        cell(ws4, rr, 4, mcap, BLUE, CUR1, border=True)
    if ev is None:
        cell(ws4, rr, 5, "n/a", BLACK, border=True, align="center")
        cell(ws4, rr, 6, rev, BLUE, CUR2, border=True)
        cell(ws4, rr, 7, "n/a", BLACK, border=True, align="center")
    else:
        cell(ws4, rr, 5, ev, BLUE, CUR1, border=True)
        cell(ws4, rr, 6, rev, BLUE, CUR2, border=True)
        cell(ws4, rr, 7, f"=IFERROR(E{rr}/F{rr},\"-\")", BLACK, MULT, border=True)
    cell(ws4, rr, 8, pe, BLUE, MULT, border=True)
    rr += 1

cell(ws4, rr + 1, 2,
     "[Note here which comps are true comps vs. reference-only, and why any obvious comp was "
     "excluded — e.g. EV/Revenue isn't meaningful for bank holding companies.]", NOTE, wrap=True)
ws4.merge_cells(start_row=rr + 1, start_column=2, end_row=rr + 1, end_column=8)
ws4.row_dimensions[rr + 1].height = 40

section_bar(ws4, rr + 3, "SCENARIO VALUATION — FORWARD REVENUE x MULTIPLE", 7)
vhdr = rr + 4
header_row(ws4, vhdr, ["", "Bear", "Base", "Bull"], start_col=2)

vr = vhdr + 1
cell(ws4, vr, 2, "Forward-year revenue ($mm) — consensus base, +/-15%", BLACK)
cell(ws4, vr, 3, "=D%d*0.85" % vr, BLACK, CUR0, border=True)
cell(ws4, vr, 4, 0, BLUE, CUR0, border=True, fill=YELLOW_FILL)
cell(ws4, vr, 5, "=D%d*1.15" % vr, BLACK, CUR0, border=True)
R_rev = vr

vr += 1
cell(ws4, vr, 2, "Assumed forward EV/Revenue multiple", BLACK)
cell(ws4, vr, 3, 0.0, BLUE, MULT, border=True, fill=YELLOW_FILL)
cell(ws4, vr, 4, 0.0, BLUE, MULT, border=True, fill=YELLOW_FILL)
cell(ws4, vr, 5, 0.0, BLUE, MULT, border=True, fill=YELLOW_FILL)
R_mult = vr

vr += 1
cell(ws4, vr, 2, "Implied enterprise value ($mm)", BLACK)
for col in ["C", "D", "E"]:
    cell(ws4, vr, ["C", "D", "E"].index(col) + 3, f"={col}{R_rev}*{col}{R_mult}", BLACK, CUR0, border=True)
R_ev = vr

vr += 1
cell(ws4, vr, 2, "Plus: net cash ($mm, cash - debt)", BLACK)
for col in ["C", "D", "E"]:
    cell(ws4, vr, ["C", "D", "E"].index(col) + 3, 0, BLUE, CUR0, border=True)
R_netcash = vr

vr += 1
cell(ws4, vr, 2, "Implied equity value ($mm)", BOLD)
for col in ["C", "D", "E"]:
    cell(ws4, vr, ["C", "D", "E"].index(col) + 3, f"={col}{R_ev}+{col}{R_netcash}", BOLD, CUR0, border=True)
R_eqval = vr

vr += 1
cell(ws4, vr, 2, "Implied share price", BOLD)
for col in ["C", "D", "E"]:
    cell(ws4, vr, ["C", "D", "E"].index(col) + 3, f"=IFERROR({col}{R_eqval}/Cover_Shares,\"-\")", BOLD, CUR2, border=True)
R_price = vr

vr += 1
cell(ws4, vr, 2, "Upside / (downside) vs. current price", BLACK)
for col in ["C", "D", "E"]:
    cell(ws4, vr, ["C", "D", "E"].index(col) + 3, f"=IFERROR({col}{R_price}/Cover_Price-1,\"-\")", BLACK, PCT0, border=True)

cell(ws4, vr + 2, 2,
     "This is a simple revenue x multiple sanity check, not a full DCF — [name the real driver "
     "model, e.g. Driver Sensitivity tab, that a rigorous valuation should feed from instead].",
     NOTE, wrap=True)
ws4.merge_cells(start_row=vr + 2, start_column=2, end_row=vr + 2, end_column=8)
ws4.row_dimensions[vr + 2].height = 40

# ============================================================================
# TAB 5 — ANALYST RATINGS
# ============================================================================
ws5 = wb.create_sheet("Analyst Ratings")
set_col_widths(ws5, [2, 20, 16, 13, 13, 13, 40, 3])
ws5.sheet_view.showGridLines = False
section_bar(ws5, 2, "SELL-SIDE COVERAGE SNAPSHOT", 6)
hdr = 4
header_row(ws5, hdr, ["Date", "Firm", "Rating", "Price Target", "Upside/(Downside)", "Notes"])
# TICKER-SPECIFIC: real ratings
ratings = [
    ("date", "Firm A", "Buy", 0),
    ("date", "Firm B", "Hold", 0),
]
rr = hdr + 1
for date, firm, rating, pt in ratings:
    cell(ws5, rr, 2, date, BLUE, border=True)
    cell(ws5, rr, 3, firm, BLUE, border=True)
    cell(ws5, rr, 4, rating, BLUE, border=True)
    if pt:
        cell(ws5, rr, 5, pt, BLUE, CUR0, border=True)
        cell(ws5, rr, 6, f"=IFERROR(E{rr}/Cover_Price-1,\"-\")", BLACK, PCT0, border=True)
    else:
        cell(ws5, rr, 5, "n/a", BLACK, border=True, align="center")
        cell(ws5, rr, 6, "-", BLACK, border=True, align="center")
    rr += 1

# ============================================================================
# TAB 6 — TECHNICAL & MARKET DATA
# ============================================================================
ws6 = wb.create_sheet("Technical & Market")
set_col_widths(ws6, [2, 34, 16, 45, 3])
ws6.sheet_view.showGridLines = False
section_bar(ws6, 2, "PRICE LEVELS, VOLATILITY, POSITIONING", 3)
hdr = 3
header_row(ws6, hdr, ["Metric", "Value", "Source / note"])
# TICKER-SPECIFIC: fill in real technical stats — price, moving averages, support/resistance,
# IV, put/call ratios, short interest, institutional/insider ownership, etc.
rr = hdr + 1
for label in ["Last price", "50-day MA", "200-day MA", "Implied volatility",
              "Short interest (% float)", "Institutional ownership"]:
    cell(ws6, rr, 2, label, BLACK, border=True)
    cell(ws6, rr, 3, 0, BLUE, border=True, align="center")
    cell(ws6, rr, 4, "source", NOTE, border=True)
    rr += 1

# ============================================================================
# TAB 7 — CATALYST CALENDAR
# ============================================================================
ws7 = wb.create_sheet("Catalyst Calendar")
set_col_widths(ws7, [2, 16, 40, 14, 3])
ws7.sheet_view.showGridLines = False
section_bar(ws7, 2, "HISTORICAL CATALYSTS", 3)
hdr = 3
header_row(ws7, hdr, ["Date", "Event", "Stock reaction"])
rr = hdr + 1
for date, ev, reaction in [("date", "[event]", "[reaction]")]:  # TICKER-SPECIFIC
    cell(ws7, rr, 2, date, BLUE, border=True)
    cell(ws7, rr, 3, ev, BLACK, border=True, wrap=True)
    cell(ws7, rr, 4, reaction, BLACK, border=True, wrap=True)
    rr += 1

rr += 1
section_bar(ws7, rr, "FORWARD CATALYSTS", 3)
rr += 1
header_row(ws7, rr, ["Date", "Event", "Why it matters"])
rr += 1
for date, ev, why in [("date", "[event]", "[why it matters]")]:  # TICKER-SPECIFIC
    cell(ws7, rr, 2, date, BLUE, border=True)
    cell(ws7, rr, 3, ev, BLACK, border=True, wrap=True)
    cell(ws7, rr, 4, why, BLACK, border=True, wrap=True)
    rr += 1

# ============================================================================
# TAB 8 — SOURCES & NOTES
# ============================================================================
ws8 = wb.create_sheet("Sources & Notes")
set_col_widths(ws8, [2, 90, 3])
ws8.sheet_view.showGridLines = False
section_bar(ws8, 2, "SOURCES & DATA-QUALITY NOTES", 1)
sources = [
    "GENERAL: research compiled [date] via secondary sources — company press releases, SEC "
    "filings, financial media, and data aggregators. Not a Bloomberg/FactSet terminal pull.",
    "",
    "[List every source URL used, grouped by topic, exactly like a bibliography.]",
    "",
    "KEY DATA-QUALITY FLAGS (do not treat as settled facts without verification):",
    "  - [List every figure that showed variance across sources, every unconfirmed date, "
    "every estimate presented elsewhere as if it were a hard fact.]",
]
rr = 4
for s in sources:
    is_header = s.isupper() or s.startswith("KEY DATA")
    cell(ws8, rr, 2, s, BOLD if is_header else NOTE, wrap=True)
    ws8.row_dimensions[rr].height = 26 if s else 8
    rr += 1

# ============================================================================
# SAVE — then run recalc.py before delivering.
# ============================================================================
OUTPUT_PATH = f"{TICKER}_Equity_Research_Model.xlsx"  # TICKER-SPECIFIC: real output filename
wb.save(OUTPUT_PATH)
print(f"Saved {OUTPUT_PATH}")
print("NEXT — both steps are blocking, do not deliver until both are clean:")
print("  1. recalculate to zero formula errors")
print(f"  2. python3 scripts/verify_model.py {OUTPUT_PATH} --price <current share price>")
