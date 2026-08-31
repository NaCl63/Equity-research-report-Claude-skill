#!/usr/bin/env python3
"""
Full 3-statement model + DCF valuation template (openpyxl).

WHEN TO USE THIS TEMPLATE vs. build_xlsx_template.py's single "Driver Sensitivity" tab:
  - A mature/steady-state company whose value depends on the INTERACTION of growth, margins,
    working capital, and capital allocation (a mega-cap tech name, an industrial, a consumer
    staple) genuinely needs a real 3-statement model — no single sensitivity grid captures it.
  - A company with one dominant, cleanly-isolable driver (a stablecoin issuer's reserve yield x
    circulation, a miner's commodity price x volume) is usually better served by the lighter
    Driver Sensitivity tab in build_xlsx_template.py — building a full 3-statement model for a
    single-driver business is overkill and invites false precision.
  - When in doubt, or for a genuinely comprehensive report, do both: the single-driver grid for
    intuition, and this full build for a defensible valuation range.

HOW TO USE THIS FILE — copy it, rename it for your ticker, and replace every block marked
"TICKER-SPECIFIC" with real, sourced numbers. Keep the structural logic as-is: it encodes three
non-obvious fixes that took real debugging to get right the first time this template was built
(for NVIDIA, August 2026) — skipping them will silently break the balance sheet:

  1. STOCK-BASED COMPENSATION MUST HIT EQUITY, NOT JUST CASH. SBC is added back to cash in the
     cash flow statement (correct — it's non-cash), but it must ALSO be added to the equity
     roll-forward (equity = prior equity + net income + SBC - buybacks - dividends), because in
     real accounting SBC simultaneously credits additional paid-in capital. Forgetting the
     equity-side add-back produces a balance-sheet gap that grows by exactly the SBC amount
     every single period — a very specific, very diagnostic symptom if you see it.
  2. ANY BALANCE-SHEET LINE THAT JUMPS BETWEEN PERIODS NEEDS AN OFFSETTING CASH-FLOW LINE. If
     debt increases (a real financing event, e.g. a bond issuance), that increase must appear as
     a financing cash inflow in the cash flow statement, or cash won't reconcile and the balance
     sheet won't balance from that period forward. The same logic applies to any other
     asset/liability line you don't hold perfectly flat.
  3. SUMMARY-LEVEL RESEARCH DATA RARELY SUMS TO THE REPORTED TOTAL. Press-release-level balance
     sheet detail (the kind you get from secondary research, not a full 10-K/10-Q) is almost
     always missing some smaller line items (prepaid expenses, deferred tax assets, lease
     liabilities, etc.). Rather than silently under-stating total assets/liabilities, add an
     explicit "reconciling plug" line sized to make your historical anchor period match the
     actually-reported total, and label it honestly as a plug — see the Balance Sheet tab below.

ALWAYS include a "balance check" row (assets minus liabilities-and-equity) on the Balance Sheet
tab, computed for every period, and verify it reads exactly 0 before delivering the model. A
clean recalc.py run does NOT catch a balance-sheet imbalance — division and arithmetic all
"work," they just don't tie out. This is the single most important verification step for this
template and is easy to skip if you only check for formula errors.

After building: run recalc.py to zero formula errors, THEN separately load the file with
data_only=True and confirm the balance-check row reads 0 for every populated period, and spot
check that the DCF's implied share price is a plausible number (not off by 1000x from a unit
mismatch, not negative, etc.) before delivering.
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import ColorScaleRule

FONT_NAME = "Arial"
BLUE = Font(name=FONT_NAME, color="0000FF", size=10)
BLACK = Font(name=FONT_NAME, color="000000", size=10)
GREEN = Font(name=FONT_NAME, color="008000", size=10)
BOLD = Font(name=FONT_NAME, bold=True, size=10)
BOLD_BLUE = Font(name=FONT_NAME, bold=True, color="0000FF", size=10)
TITLE = Font(name=FONT_NAME, bold=True, size=16, color="1F3864")
SUBTITLE = Font(name=FONT_NAME, size=11, italic=True, color="595959")
HDR = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
SECTION = Font(name=FONT_NAME, bold=True, size=12, color="FFFFFF")
NOTE = Font(name=FONT_NAME, italic=True, size=9, color="595959")
ITAL_RED = Font(name=FONT_NAME, italic=True, size=8, color="B22222")

HDR_FILL = PatternFill("solid", fgColor="1F3864")
SECTION_FILL = PatternFill("solid", fgColor="2E5395")
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
LIGHT_FILL = PatternFill("solid", fgColor="D9E2F3")
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
HIST_FILL = PatternFill("solid", fgColor="E7E6E6")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CUR0 = '$#,##0;($#,##0);"-"'
CUR1 = '$#,##0.0;($#,##0.0);"-"'
CUR2 = '$#,##0.00;($#,##0.00);"-"'
PCT1 = '0.0%;(0.0%);"-"'
PCT0 = '0%;(0%);"-"'
MULT = '0.00"x"'
NUM0 = '#,##0;(#,##0);"-"'
DAYS0 = '0.0"d"'

wb = Workbook()

PERIODS = ["FY2024A", "FY2025A", "FY2026A", "FY2027E", "FY2028E", "FY2029E", "FY2030E", "FY2031E"]
HIST_N = 3  # first 3 columns are historical/actual
COLS = [get_column_letter(3 + i) for i in range(len(PERIODS))]  # C..J


def section_bar(ws, row, text, span, start_col=1):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + span - 1)
    c = ws.cell(row=row, column=start_col, value=text)
    c.font = SECTION
    c.fill = SECTION_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def cell(ws, row, col, value, font=BLACK, fmt=None, fill=None, align=None, border=False, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if align or wrap:
        c.alignment = Alignment(horizontal=align or "general", wrap_text=wrap, vertical="center")
    if border:
        c.border = BORDER
    return c


def period_header(ws, row, start_col=3):
    for i, p in enumerate(PERIODS):
        fill = HIST_FILL if i < HIST_N else LIGHT_FILL
        c = cell(ws, row, start_col + i, p, HDR if i < HIST_N else BOLD_BLUE, fill=HDR_FILL if i < HIST_N else LIGHT_FILL,
                 align="center", border=True)
    ws.row_dimensions[row].height = 16


def data_row(ws, row, label, values, fmt=CUR0, bold=False, hist_font=BLUE, formula=False, note=None, note_col=11):
    """values: list of 8 items (per PERIODS) — numbers (hardcoded, historical or assumption) or
    formula strings (start with '='). None = leave blank."""
    cell(ws, row, 2, label, BOLD if bold else BLACK)
    for i, v in enumerate(values):
        col = 3 + i
        if v is None:
            continue
        is_formula = isinstance(v, str) and v.startswith("=")
        f = BLACK if is_formula else (hist_font if i < HIST_N else BLUE)
        fill = YELLOW_FILL if (not is_formula and i >= HIST_N) else None
        cell(ws, row, col, v, f, fmt, fill=fill, border=True)
    if note:
        cell(ws, row, note_col, note, NOTE, wrap=True)
    return row


# ============================================================================
# TAB 1 — COVER
# ============================================================================
ws = wb.active
ws.title = "Cover"
set_col_widths(ws, [3, 30, 18, 18, 18, 18, 3])
ws.sheet_view.showGridLines = False

# TICKER-SPECIFIC: replace with the real company name / ticker / model date / fiscal year note
COMPANY_NAME = "COMPANY NAME, INC."
TICKER = "TICK"
EXCHANGE = "NASDAQ"
MODEL_DATE = "YYYY-MM-DD"
cell(ws, 2, 2, COMPANY_NAME, TITLE)
cell(ws, 3, 2, f"{EXCHANGE}: {TICKER} — 3-Statement Model + DCF Valuation", SUBTITLE)
cell(ws, 4, 2, "Illustrative financial model for educational purposes — not investment advice", NOTE)
cell(ws, 5, 2, "Model date:", BOLD); cell(ws, 5, 3, MODEL_DATE, BLUE)
cell(ws, 6, 2, "Fiscal year end:", BOLD); cell(ws, 6, 3, "[e.g. Dec 31 / late Jan — state actual FYE and most recent close date]", NOTE)

# TICKER-SPECIFIC: replace every value below with real, sourced market data. Keep the row order —
# other tabs reference these rows by the price_row/shares_row/debt_row/cash_row variables computed
# in the loop just below, not by hardcoded coordinates.
section_bar(ws, 8, "MARKET SNAPSHOT (as of [date])", 5)
snap = [
    ("Share price", 100.00, CUR2, "[source, date of close]"),
    ("52-week high / low", "$000.00 / $000.00", None, "[source]"),
    ("Diluted shares outstanding (mm)", 1000, NUM0, "[source, ~date]"),
    ("Market capitalization ($mm)", None, CUR0, "Formula: price x shares out"),
    ("Total debt ($mm)", 1000, CUR0, "[source — note any recent financing events that could be stale]"),
    ("Cash + short-term investments ($mm)", 1000, CUR0, "[source]"),
    ("Net cash ($mm)", None, CUR0, "Formula: cash - debt"),
    ("Beta (5-year)", 1.00, "0.00", "[source — cross-check vs. a second data provider, methodologies differ]"),
]
price_row = 9
r = price_row
for label, val, fmt, src in snap:
    cell(ws, r, 2, label, BLACK)
    if label.startswith("Market cap"):
        cell(ws, r, 3, "=Cover_Price*Cover_Shares", BLACK, CUR0)
        mcap_row = r
    elif label.startswith("Net cash"):
        cell(ws, r, 3, f"=C{cash_row}-C{debt_row}", BLACK, CUR0)
    elif isinstance(val, str):
        cell(ws, r, 3, val, BLUE)
    else:
        cell(ws, r, 3, val, BLUE, fmt)
        if label.startswith("Diluted shares"):
            shares_row = r
        if label.startswith("Total debt"):
            debt_row = r
        if label.startswith("Cash + short"):
            cash_row = r
    cell(ws, r, 5, src, NOTE)
    r += 1

wb.defined_names["Cover_Price"] = DefinedName("Cover_Price", attr_text=f"Cover!$C${price_row}")
wb.defined_names["Cover_Shares"] = DefinedName("Cover_Shares", attr_text=f"Cover!$C${shares_row}")
wb.defined_names["Cover_MktCap"] = DefinedName("Cover_MktCap", attr_text=f"Cover!$C${mcap_row}")
wb.defined_names["Cover_Debt"] = DefinedName("Cover_Debt", attr_text=f"Cover!$C${debt_row}")
wb.defined_names["Cover_Cash"] = DefinedName("Cover_Cash", attr_text=f"Cover!$C${cash_row}")

# TICKER-SPECIFIC: rewrite this paragraph to name the real reasons THIS company's valuation
# depends on multiple interacting drivers (not just growth/margin/capital-return in the abstract),
# and to name the real, specific uncertainties around the forecast (not a generic risk list).
section_bar(ws, r + 1, "WHAT THIS MODEL IS AND ISN'T", 5)
cell(ws, r + 2, 2,
     "This is a full 3-statement model (income statement, balance sheet, cash flow) with a "
     "5-year explicit forecast built off historical actuals, feeding a DCF valuation. It goes "
     "well beyond a single-driver sensitivity model because [COMPANY]'s valuation genuinely "
     "depends on the interaction of growth, margin trajectory, and capital return, not one lever "
     "— [explain why]. That said: every projected year is a scenario built on explicit, visible "
     "assumptions (yellow cells) — not a prediction. [Name the real, specific sources of "
     "uncertainty for this company/sector — see Sources tab]. Change the yellow assumption cells "
     "to build your own scenario.",
     NOTE, wrap=True)
ws.row_dimensions[r + 2].height = 90
ws.merge_cells(start_row=r + 2, start_column=2, end_row=r + 2, end_column=6)

cell(ws, r + 4, 2, "Tabs: Income Statement | Balance Sheet | Cash Flow & Working Capital | DCF & WACC | Sensitivity | Sources & Assumptions", NOTE, wrap=True)
ws.merge_cells(start_row=r + 4, start_column=2, end_row=r + 4, end_column=6)

# ============================================================================
# TAB 2 — INCOME STATEMENT
# ============================================================================
ws2 = wb.create_sheet("Income Statement")
set_col_widths(ws2, [2, 32] + [12] * 8 + [3])
ws2.sheet_view.showGridLines = False
section_bar(ws2, 2, "INCOME STATEMENT — $mm except per-share", 10)
cell(ws2, 3, 2, "Gray columns = historical actuals (10-K/press releases). Blue/yellow columns = "
     "explicit forecast assumptions — change any yellow cell to rebuild the scenario.", NOTE, wrap=True)
ws2.merge_cells(start_row=3, start_column=2, end_row=3, end_column=10)
hdr_row = 5
cell(ws2, hdr_row, 2, "", BLACK)
period_header(ws2, hdr_row)
r = hdr_row + 1

# --- Revenue build ---
# TICKER-SPECIFIC: replace the three historical revenue figures with real reported actuals, and
# set FY[+1]E-FY[+5]E growth assumptions. Ground the first forecast year in the hardest data you
# have (most recent quarter actual + next-quarter guidance, annualized) rather than picking a
# round number — that's the single highest-value sanity check in the whole model, and catching an
# assumption that's inconsistent with quarters already reported is exactly the kind of error this
# note exists to prevent.
R_rev = r
data_row(ws2, r, "Total revenue", [10000, 12000, 14000, None, None, None, None, None], CUR0, bold=True,
          note="[Y1]-[Y3] actual per 10-K/press releases — replace with real figures")
r += 1
R_revgrowth = r
data_row(ws2, r, "  Revenue growth YoY", [None, "=D%d/C%d-1" % (R_rev, R_rev), "=E%d/D%d-1" % (R_rev, R_rev),
                                          0.15, 0.13, 0.11, 0.09, 0.08], PCT1,
         note="Forecast-year growth is an explicit, editable assumption, not a forecast. "
              "Calibrate the first forecast year against the hardest data available (most recent "
              "actual quarter + next-quarter guidance, annualized) rather than an arbitrary round "
              "number — cross-check any Street-consensus figure you use the same way before "
              "trusting it. Later years should decelerate/stabilize based on explicit judgment "
              "documented here, not silently — edit freely.")
r += 1
for i in range(HIST_N, len(PERIODS)):
    col = COLS[i]
    prev_col = COLS[i - 1]
    ws2.cell(row=R_rev, column=3 + i).value = f"={prev_col}{R_rev}*(1+{col}{R_revgrowth})"
    ws2.cell(row=R_rev, column=3 + i).font = BLACK
    ws2.cell(row=R_rev, column=3 + i).number_format = CUR0

r += 1
R_gm = r
# TICKER-SPECIFIC: replace the three hardcoded COGS figures below with real reported COGS
data_row(ws2, r, "Gross margin %", [
    "=1-6500/C%d" % R_rev, "=1-7680/D%d" % R_rev, "=1-8680/E%d" % R_rev,
    0.380, 0.385, 0.390, 0.390, 0.385], PCT1,
    note="[Y1]-[Y3] derived from reported COGS. Forecast years: state the real thesis for the "
         "margin trajectory (mix shift, input costs, competitive pricing, scale economies) rather "
         "than holding flat by default — a flat assumption should be a deliberate choice, not the "
         "absence of one.")
r += 1
R_gp = r
data_row(ws2, r, "Gross profit", [f"={c}{R_rev}*{c}{R_gm}" for c in COLS], CUR0, bold=True, formula=True)
r += 1
R_rdpct = r
# TICKER-SPECIFIC: replace the three hardcoded R&D $ figures with real reported R&D expense
data_row(ws2, r, "R&D % of revenue", [
    "=850/C%d" % R_rev, "=960/D%d" % R_rev, "=1050/E%d" % R_rev,
    0.075, 0.075, 0.075, 0.075, 0.075], PCT1,
    note="[State the real thesis for R&D intensity trend — e.g. investing ahead of a new product "
         "cycle, or scaling down as the platform matures] — adjust to taste.")
r += 1
R_rd = r
data_row(ws2, r, "R&D expense", [f"={c}{R_rev}*{c}{R_rdpct}" for c in COLS], CUR0, formula=True)
r += 1
R_sgapct = r
# TICKER-SPECIFIC: replace the three hardcoded SG&A $ figures with real reported SG&A expense
data_row(ws2, r, "SG&A % of revenue", [
    "=1200/C%d" % R_rev, "=1350/D%d" % R_rev, "=1500/E%d" % R_rev,
    0.105, 0.103, 0.100, 0.098, 0.095], PCT1)
r += 1
R_sga = r
data_row(ws2, r, "SG&A expense", [f"={c}{R_rev}*{c}{R_sgapct}" for c in COLS], CUR0, formula=True)
r += 1
R_ebit = r
data_row(ws2, r, "Operating income (EBIT)", [f"={c}{R_gp}-{c}{R_rd}-{c}{R_sga}" for c in COLS], CUR0, bold=True, formula=True)
r += 1
R_ebitmargin = r
data_row(ws2, r, "  Operating margin %", [f"=IFERROR({c}{R_ebit}/{c}{R_rev},\"-\")" for c in COLS], PCT1, formula=True)
r += 1

r += 1
# TICKER-SPECIFIC: if this company has large, volatile non-operating items (equity-securities
# mark-to-market gains, FX, litigation settlements) that would distort a forward operating model
# if naively extrapolated, exclude them here and say so explicitly — as done below. If not
# applicable, delete this note.
cell(ws2, r, 2, "Non-operating items (exclude anything large/volatile/non-recurring — see note)", NOTE, wrap=True)
r += 1
R_cash_yield = r
# TICKER-SPECIFIC: the actual yield the company earns on its cash/investments portfolio. Ground it
# in the disclosed interest income / average cash balance for the last reported year, then fade it
# toward your forward short-rate view (FRED DGS3MO / DGS1MO — see references/sec_edgar_fred.md).
data_row(ws2, r, "  Yield on cash & investments (assumption)", [None, None, 0.045, 0.042, 0.038, 0.035, 0.033, 0.032], PCT1,
         note="Drives net interest income below off the PRIOR period's ending cash + short-term "
              "investments. Anchor it to disclosed interest income / average cash for the last "
              "reported year rather than a generic short rate.")
r += 1
R_debt_rate = r
# TICKER-SPECIFIC: blended coupon on the company's actual debt stack, from the 10-K debt footnote.
data_row(ws2, r, "  Blended interest rate on debt (assumption)", [None, None, 0.055, 0.055, 0.055, 0.055, 0.055, 0.055], PCT1,
         note="Blended coupon from the 10-K long-term debt footnote. Should be consistent with "
              "the pre-tax cost of debt used in the WACC build on the DCF tab.")
r += 1
R_int = r
# Historical columns are hardcoded actuals; forecast columns are REWIRED to a formula further down
# (see "wire the Income Statement's interest and share count" block) once the Balance Sheet rows
# exist. Interest is computed off the PRIOR period's ending balances on purpose — using the same
# period's closing cash would make cash depend on interest which depends on cash, a circular
# reference this toolchain's recalc step cannot resolve. Beginning-balance convention is standard
# practice and avoids the circularity outright rather than papering over it with iteration.
data_row(ws2, r, "Net interest income", [50, 60, 70, None, None, None, None, None], CUR0,
         note="[Y1]-[Y3] = interest income less interest expense, actual. Forecast years = prior "
              "period ending (cash + short-term investments) x cash yield, less prior period "
              "ending debt x blended debt rate — formula-linked to the Balance Sheet, so a change "
              "in the cash build or a financing event flows through to earnings automatically.")
r += 1
R_tax_rate = r
data_row(ws2, r, "Effective tax rate", [0.180, 0.190, 0.195, 0.200, 0.205, 0.210, 0.210, 0.210], PCT1,
         note="[Y1]-[Y3] actual. Forecast years: step toward the statutory rate over the explicit "
              "forecast unless the company has a specific, disclosed reason to expect otherwise "
              "(tax credits, foreign mix, etc.) — state the reasoning here.")
r += 1
R_pretax = r
data_row(ws2, r, "Pretax income (op. income + net interest)", [f"={c}{R_ebit}+{c}{R_int}" for c in COLS], CUR0, formula=True)
r += 1
R_taxexp = r
data_row(ws2, r, "Income tax expense", [f"={c}{R_pretax}*{c}{R_tax_rate}" for c in COLS], CUR0, formula=True)
r += 1
R_ni = r
# TICKER-SPECIFIC: rename this row and its note if the company has no material non-operating
# items to exclude — "Net income" is fine as-is in that case.
data_row(ws2, r, "Net income (excl. [name any excluded non-operating items])", [f"={c}{R_pretax}-{c}{R_taxexp}" for c in COLS], CUR0, bold=True, formula=True,
         note="[If applicable: name and size the excluded non-operating item(s) here, with "
              "dollar figures and why they'd distort a forward operating model if extrapolated. "
              "Delete this note if there's nothing material to exclude.]")
r += 1
R_sbc_dil_factor = r
# TICKER-SPECIFIC: fraction of gross SBC dollars that actually becomes NEW shares. Most companies
# net-settle RSUs for tax withholding, so gross issuance is materially below SBC$/price — pull the
# real gross-issued and withheld share counts from the 10-K statement of stockholders' equity and
# back the factor out of those, rather than leaving the conservative 1.00 default.
data_row(ws2, r, "  SBC-to-share-issuance factor (net of withholding)", [None, None, None, 1.00, 1.00, 1.00, 1.00, 1.00], "0.00",
         note="1.00 = every SBC dollar issues shares at the current price (maximum dilution, "
              "deliberately conservative). Replace with (shares issued under equity plans) / "
              "(SBC$ / avg price) from the 10-K equity statement.")
r += 1
R_shares = r
# Historical columns are hardcoded actuals; forecast columns are REWIRED below to
#   prior shares + SBC-driven issuance - buyback-retired shares
# so that the share count, the buyback line on the cash flow statement, and the SBC add-back in the
# equity roll-forward all move together instead of being three independently-typed assumptions.
data_row(ws2, r, "Diluted shares outstanding (mm)", [1020, 1010, 1000, None, None, None, None, None], NUM0,
         note="[Y1]-[Y3] actual per 10-K. Forecast years = prior count + (SBC x issuance factor "
              "/ share price) - (buybacks / share price). Both legs are held at the CURRENT share "
              "price, which understates retirement if the stock rises and overstates it if it "
              "falls — state that assumption in the report rather than implying precision.")
r += 1
R_eps = r
data_row(ws2, r, "Diluted EPS (excl. equity-securities gains)", [f"=IFERROR({c}{R_ni}/{c}{R_shares},\"-\")" for c in COLS], CUR2, bold=True, formula=True)

for rr_ in range(hdr_row + 1, r + 1):
    for cc in range(2, 11):
        ws2.cell(row=rr_, column=cc).border = BORDER

# ============================================================================
# TAB 3 — BALANCE SHEET
# ============================================================================
ws3 = wb.create_sheet("Balance Sheet")
set_col_widths(ws3, [2, 32] + [12] * 8 + [3])
ws3.sheet_view.showGridLines = False
section_bar(ws3, 2, "BALANCE SHEET — $mm", 10)
# TICKER-SPECIFIC: update the fiscal-year-end date reference below
cell(ws3, 3, 2, "Only the most recent fiscal year-end shown as the historical anchor for "
     "simplicity — a full model would carry 3 years, but the projection only needs one clean "
     "starting balance sheet to roll forward from. Working capital and PP&E roll forward using "
     "the schedules on the Cash Flow tab.", NOTE, wrap=True)
ws3.merge_cells(start_row=3, start_column=2, end_row=3, end_column=10)
ws3.row_dimensions[3].height = 28
hdr_row = 6
cell(ws3, hdr_row, 2, "", BLACK)
period_header(ws3, hdr_row)
r = hdr_row + 1

# TICKER-SPECIFIC: replace every hardcoded value below with the real historical anchor-period
# balance sheet. Keep the row structure — Cash/AR/Inventory/PP&E projected cells get overwritten
# by formulas further down (linked to the Cash Flow tab schedules), so their forecast-column
# values here should stay None.
R_cash = r
data_row(ws3, r, "Cash & equivalents", [None, None, 500] + [None] * 5, CUR0, note="Plug — computed on Cash Flow tab, linked here")
r += 1
R_sti = r
data_row(ws3, r, "Short-term investments", [None, None, 300, 300, 300, 300, 300, 300], CUR0,
         note="Held flat for simplicity — a fuller model would flex this with excess-cash policy")
r += 1
R_ar = r
data_row(ws3, r, "Accounts receivable", [None, None, 900] + [None] * 5, CUR0, note="Projected via DSO on Cash Flow tab")
r += 1
R_inv = r
data_row(ws3, r, "Inventory", [None, None, 700] + [None] * 5, CUR0, note="Projected via DIO on Cash Flow tab")
r += 1
R_ppe = r
data_row(ws3, r, "PP&E, net", [None, None, 1200] + [None] * 5, CUR0, note="Rolls forward: prior + capex - D&A (Cash Flow tab)")
r += 1
R_gwintang = r
data_row(ws3, r, "Goodwill & intangibles", [None, None, 800, 800, 800, 800, 800, 800], CUR0, note="Held flat — no M&A assumed")
r += 1
R_othassets = r
# TICKER-SPECIFIC: rename/resize this line for whatever large "other" asset category this
# company actually carries (equity-method investments, deferred tax assets, right-of-use assets,
# etc.) — or delete the row if not applicable. Held flat by design (see note pattern below).
data_row(ws3, r, "Other assets (name the real category)", [None, None, 400, 400, 400, 400, 400, 400], CUR0,
         note="If this line is large and genuinely unpredictable, hold it flat rather than "
              "guessing a growth rate with no cash-flow treatment — any change here needs an "
              "offsetting cash-flow line to keep the balance sheet tying out.")
r += 1
R_assetplug = r
# TICKER-SPECIFIC: size this plug = actual reported total assets (anchor period) minus the sum
# of the itemized asset lines above. This is very likely necessary — press-release-level detail
# almost never sums to the full 10-K total. Label it honestly; don't skip it and don't hide it.
data_row(ws3, r, "Other assets (reconciling plug)", [None, None, 0, 0, 0, 0, 0, 0], CUR0,
         note="Anchor-period value = actual reported total assets minus this model's itemized "
              "asset lines above. If the itemized press-release-level detail this model was "
              "built from doesn't sum to the reported total (missing prepaid expenses, deferred "
              "tax assets, lease right-of-use assets, etc.), size this plug to close that gap "
              "explicitly rather than silently under-stating total assets. Replace with the "
              "fully itemized 10-K balance sheet for production use; held flat here as a "
              "simplification.")
r += 1
R_totassets = r
data_row(ws3, r, "Total assets", [None, None] + [f"=SUM({c}{R_cash}:{c}{R_assetplug})" for c in COLS[2:]], CUR0, bold=True, formula=True)
r += 2

R_ap = r
data_row(ws3, r, "Accounts payable", [None, None, 350] + [None] * 5, CUR0, note="Projected via DPO on Cash Flow tab")
r += 1
R_otherliab_curr = r
data_row(ws3, r, "Other current liabilities", [None, None, 450, 450, 450, 450, 450, 450], CUR0, note="Held flat as % structure — simplification")
r += 1
R_debt = r
# TICKER-SPECIFIC: replace with real debt figures. If a financing event (bond issuance, term
# loan, repayment) causes debt to jump between periods, make sure the offsetting cash-flow line
# on the Cash Flow tab (see R_debtissuance below) is wired to this row — it already is by formula,
# just confirm the jump you enter here is real and dated.
data_row(ws3, r, "Long-term debt", [None, None, 600, 600, 600, 600, 600, 600], CUR0,
         note="[State any known/planned financing events and dates that would change this line]")
r += 1
R_otherliab_lt = r
data_row(ws3, r, "Other long-term liabilities", [None, None, 300, 300, 300, 300, 300, 300], CUR0, note="Held flat")
r += 1
R_liabplug = r
# TICKER-SPECIFIC: size this plug = actual reported total liabilities (anchor period) minus the
# sum of the itemized liability lines above — same logic as the asset-side plug.
data_row(ws3, r, "Other liabilities (reconciling plug)", [None, None, 0, 0, 0, 0, 0, 0], CUR0,
         note="Same treatment as the asset-side plug above: anchor-period value = actual "
              "reported total liabilities minus this model's itemized liability lines (likely "
              "income tax payable, operating lease liabilities, and other items not broken out "
              "at press-release-level detail). Held flat as a simplification.")
r += 1
R_totliab = r
data_row(ws3, r, "Total liabilities", [None, None] + [f"=SUM({c}{R_ap}:{c}{R_liabplug})" for c in COLS[2:]], CUR0, bold=True, formula=True)
r += 1
R_equity_open = r
data_row(ws3, r, "Stockholders' equity, beginning of period", [None, None, None] + [None] * 5, CUR0,
         note="First forecast year = anchor-period ending equity; thereafter = prior year ending equity")
r += 1
R_equity_close = r
# TICKER-SPECIFIC: replace with real anchor-period reported total stockholders' equity
data_row(ws3, r, "Stockholders' equity, end of period", [None, None, 3100] + [None] * 5, CUR0, bold=True,
         note="Anchor period actual. Forecast years = beginning equity + net income + SBC - "
              "buybacks - dividends (Cash Flow tab) — SBC must be added here, not just in cash, "
              "see module docstring fix #1.")
r += 1
R_totliabeq = r
data_row(ws3, r, "Total liabilities & equity", [None, None] + [f"={c}{R_totliab}+{c}{R_equity_close}" for c in COLS[2:]], CUR0, bold=True, formula=True)
r += 1
R_balcheck = r
data_row(ws3, r, "Balance check (assets - liab&eq, should be 0)", [None, None, None, None, None, None, None, None], CUR0,
         note="Formula added below for FY26A-FY31E — should read exactly 0 in every populated column")

for rr_ in range(hdr_row + 1, r + 1):
    for cc in range(2, 11):
        ws3.cell(row=rr_, column=cc).border = BORDER

# ============================================================================
# TAB 4 — CASH FLOW & WORKING CAPITAL (this is where the 3 statements actually link up)
# ============================================================================
ws4 = wb.create_sheet("Cash Flow & Working Capital")
set_col_widths(ws4, [2, 34] + [12] * 8 + [3])
ws4.sheet_view.showGridLines = False
section_bar(ws4, 2, "WORKING CAPITAL SCHEDULE (drives AR/Inventory/AP on the Balance Sheet)", 10)
hdr_row = 3
cell(ws4, hdr_row, 2, "", BLACK)
period_header(ws4, hdr_row)
r = hdr_row + 1

# TICKER-SPECIFIC: the anchor-period (column E / index 2) DSO/DIO/DPO formulas below reference
# hardcoded AR/Inventory/AP/COGS dollar figures — replace 900/350 (AR/AP) and 700/COGS with the
# real anchor-period figures matching what you entered on the Balance Sheet tab, then set
# forecast-year day-counts based on the company's actual trend (improving, stable, deteriorating).
R_dso = r
data_row(ws4, r, "Days sales outstanding (AR/Rev x 365)",
         [None, None, None, 55, 55, 54, 54, 53], DAYS0,
         note="Anchor period = real AR/Revenue*365 (formula, not hardcoded) — replace the anchor "
              "cell formula's inputs with real figures matching the Balance Sheet tab.")
ws4.cell(row=R_dso, column=5).value = "=900/14000*365"
r += 1
R_dio = r
data_row(ws4, r, "Days inventory outstanding (Inv/COGS x 365)",
         [None, None, None, 45, 44, 43, 42, 41], DAYS0)
ws4.cell(row=R_dio, column=5).value = "=700/8680*365"
r += 1
R_dpo = r
data_row(ws4, r, "Days payable outstanding (AP/COGS x 365)",
         [None, None, None, 40, 40, 40, 40, 40], DAYS0)
ws4.cell(row=R_dpo, column=5).value = "=350/8680*365"
r += 2

section_bar(ws4, r, "CAPEX & DEPRECIATION SCHEDULE (drives PP&E on the Balance Sheet)", 10)
r += 1
hdr2 = r
period_header(ws4, hdr2)
r += 1
R_capexpct = r
# TICKER-SPECIFIC: replace 700/14000 with real anchor-period capex/revenue; note any structural
# reason capex intensity differs from peers (asset-light/fabless, capital-intensive build-out, etc.)
data_row(ws4, r, "Capex % of revenue", [None, None, "=700/14000", 0.050, 0.052, 0.050, 0.048, 0.045], PCT1,
         note="[State the real capital-intensity story for this business]")
r += 1
R_capex = r
data_row(ws4, r, "Capital expenditures", [None, None, 700] + [f"={c}Rev*{c}{R_capexpct}" for c in COLS[3:]], CUR0, formula=True)
# fix formula refs to point at income statement revenue row properly
for i, c in enumerate(COLS[3:], start=3):
    ws4.cell(row=R_capex, column=3 + i).value = f"='Income Statement'!{c}{R_rev}*{c}{R_capexpct}"
r += 1
R_da = r
# TICKER-SPECIFIC: replace with real D&A figures
data_row(ws4, r, "Depreciation & amortization", [None, None, 250, 280, 310, 340, 370, 400], CUR0,
         note="Anchor period actual; forecast years illustrative growth roughly tracking the "
              "expanding PP&E base")

for rr_ in range(hdr_row + 1, r + 1):
    for cc in range(2, 11):
        ws4.cell(row=rr_, column=cc).border = BORDER

r += 2
section_bar(ws4, r, "CASH FLOW STATEMENT (derived — links Income Statement to Balance Sheet)", 10)
r += 1
hdr3 = r
period_header(ws4, hdr3)
r += 1
R_ni_cf = r
# TICKER-SPECIFIC: replace the hardcoded anchor-period net income figure (must match the real
# reported anchor-period net income, not necessarily equal to the formula-driven forecast years'
# "excl. non-operating items" definition if those differ)
data_row(ws4, r, "Net income", [None, None, 1600] + [f"='Income Statement'!{c}{R_ni}" for c in COLS[3:]], CUR0, formula=True)
r += 1
R_da_cf = r
data_row(ws4, r, "+ D&A", [None, None] + [f"={c}{R_da}" for c in COLS[2:]], CUR0, formula=True)
r += 1
R_sbc = r
# TICKER-SPECIFIC: replace with real SBC figures
data_row(ws4, r, "+ Stock-based compensation", [None, None, 400, 440, 480, 520, 560, 600], CUR0,
         note="Anchor period actual; forecast years illustrative — continues growing with "
              "headcount/revenue scale")
r += 1
R_dwc = r
# TICKER-SPECIFIC: replace the hardcoded anchor-period NWC change with the real figure per the
# reported cash flow statement (this is a check value only — forecast years are pure formula)
data_row(ws4, r, "- Increase in net working capital", [None, None, -120] + [None] * 5, CUR0,
         note="Anchor period actual (per reported cash flow statement). Forecast years computed "
              "from the DSO/DIO/DPO schedule above vs. the prior period's implied AR/Inventory/AP "
              "— see formulas, do not hardcode these.")
for i, c in enumerate(COLS[3:], start=3):
    prev_c = COLS[i - 1]
    # implied AR/Inv/AP this period vs prior period, using DSO/DIO/DPO and IS revenue/COGS
    f = (f"=-((({c}{R_dso}/365)*'Income Statement'!{c}{R_rev}-({prev_c}{R_dso}/365)*'Income Statement'!{prev_c}{R_rev})"
         f"+(({c}{R_dio}/365)*('Income Statement'!{c}{R_rev}*(1-'Income Statement'!{c}{R_gm}))-({prev_c}{R_dio}/365)*('Income Statement'!{prev_c}{R_rev}*(1-'Income Statement'!{prev_c}{R_gm})))"
         f"-(({c}{R_dpo}/365)*('Income Statement'!{c}{R_rev}*(1-'Income Statement'!{c}{R_gm}))-({prev_c}{R_dpo}/365)*('Income Statement'!{prev_c}{R_rev}*(1-'Income Statement'!{prev_c}{R_gm}))))")
    ws4.cell(row=R_dwc, column=3 + i).value = f
    ws4.cell(row=R_dwc, column=3 + i).font = BLACK
    ws4.cell(row=R_dwc, column=3 + i).number_format = CUR0
    ws4.cell(row=R_dwc, column=3 + i).border = BORDER
r += 1
R_cfo = r
data_row(ws4, r, "Cash from operations", [None, None] + [f"={c}{R_ni_cf}+{c}{R_da_cf}+{c}{R_sbc}+{c}{R_dwc}" for c in COLS[2:]], CUR0, bold=True, formula=True)
r += 1
R_capex_cf = r
data_row(ws4, r, "- Capital expenditures", [None, None] + [f"=-{c}{R_capex}" for c in COLS[2:]], CUR0, formula=True)
r += 1
R_fcf = r
data_row(ws4, r, "Free cash flow (levered, after SBC add-back)", [None, None] + [f"={c}{R_cfo}+{c}{R_capex_cf}" for c in COLS[2:]], CUR0, bold=True, formula=True,
         note="NOT the same figure as 'Unlevered free cash flow' on the DCF tab, and the gap is "
              "intentional. This line is CFO - capex: it adds SBC back as non-cash and is stated "
              "after interest and tax as actually paid. The DCF's UFCF starts from EBIT, taxes it "
              "at the effective rate, and does NOT add SBC back — SBC is a real economic cost of "
              "compensating employees, and adding it back while ALSO diluting the share count "
              "would double-count the benefit. Expect the DCF figure to be lower by roughly the "
              "SBC amount; if anyone asks why the two differ, that is the answer.")
r += 1
R_buyback = r
# TICKER-SPECIFIC: replace with real buyback figures and cite the real authorization size/date
data_row(ws4, r, "- Share buybacks", [None, None, -200, -220, -240, -250, -250, -250], CUR0,
         note="Anchor period actual. Forecast years: [cite the real buyback authorization "
              "size/date this pace is based on] — adjust to your own capital-return assumption")
r += 1
R_div = r
# TICKER-SPECIFIC: replace with real dividend figures, or zero out entirely if the company pays none
data_row(ws4, r, "- Dividends paid", [None, None, -80, -85, -90, -95, -100, -105], CUR0,
         note="Anchor period actual. Forecast years: [cite the real dividend per share and "
              "declaration date this is based on, applied to the declining/growing share count]")
r += 1
R_debtissuance = r
data_row(ws4, r, "+ Net debt issuance / (repayment)", [None] * 8, CUR0,
         note="Formula: change in Balance Sheet long-term debt period-over-period, so the "
              "FY27E jump to the post-bond-deal ~$12.8bn debt level shows up as a financing "
              "cash inflow here rather than silently breaking the balance sheet.")
for i, c in enumerate(COLS[3:], start=3):
    prev_c = COLS[i - 1]
    ws4.cell(row=R_debtissuance, column=3 + i).value = f"='Balance Sheet'!{c}{R_debt}-'Balance Sheet'!{prev_c}{R_debt}"
    ws4.cell(row=R_debtissuance, column=3 + i).font = BLACK
    ws4.cell(row=R_debtissuance, column=3 + i).number_format = CUR0
    ws4.cell(row=R_debtissuance, column=3 + i).border = BORDER
r += 1
R_dcash = r
data_row(ws4, r, "Net change in cash", [None, None] + [f"={c}{R_fcf}+{c}{R_buyback}+{c}{R_div}+{c}{R_debtissuance}" for c in COLS[2:]], CUR0, bold=True, formula=True)

for rr_ in range(hdr3, r + 1):
    for cc in range(2, 11):
        ws4.cell(row=rr_, column=cc).border = BORDER

# ---- Now wire the Balance Sheet projected cells back to these schedules ----
for i, c in enumerate(COLS[3:], start=3):
    prev_c = COLS[i - 1]
    # Cash: prior cash + net change in cash
    ws3.cell(row=R_cash, column=3 + i).value = f"={prev_c}{R_cash}+'Cash Flow & Working Capital'!{c}{R_dcash}"
    ws3.cell(row=R_cash, column=3 + i).font = BLACK
    ws3.cell(row=R_cash, column=3 + i).number_format = CUR0
    ws3.cell(row=R_cash, column=3 + i).border = BORDER
    # AR = DSO/365 * revenue
    ws3.cell(row=R_ar, column=3 + i).value = f"=('Cash Flow & Working Capital'!{c}{R_dso}/365)*'Income Statement'!{c}{R_rev}"
    ws3.cell(row=R_ar, column=3 + i).font = BLACK
    ws3.cell(row=R_ar, column=3 + i).number_format = CUR0
    ws3.cell(row=R_ar, column=3 + i).border = BORDER
    # Inventory = DIO/365 * COGS ; COGS = revenue*(1-gross margin)
    ws3.cell(row=R_inv, column=3 + i).value = (
        f"=('Cash Flow & Working Capital'!{c}{R_dio}/365)*('Income Statement'!{c}{R_rev}*(1-'Income Statement'!{c}{R_gm}))"
    )
    ws3.cell(row=R_inv, column=3 + i).font = BLACK
    ws3.cell(row=R_inv, column=3 + i).number_format = CUR0
    ws3.cell(row=R_inv, column=3 + i).border = BORDER
    # AP = DPO/365 * COGS
    ws3.cell(row=R_ap, column=3 + i).value = (
        f"=('Cash Flow & Working Capital'!{c}{R_dpo}/365)*('Income Statement'!{c}{R_rev}*(1-'Income Statement'!{c}{R_gm}))"
    )
    ws3.cell(row=R_ap, column=3 + i).font = BLACK
    ws3.cell(row=R_ap, column=3 + i).number_format = CUR0
    ws3.cell(row=R_ap, column=3 + i).border = BORDER
    # PP&E = prior PP&E + capex - D&A
    ws3.cell(row=R_ppe, column=3 + i).value = f"={prev_c}{R_ppe}+'Cash Flow & Working Capital'!{c}{R_capex}-'Cash Flow & Working Capital'!{c}{R_da}"
    ws3.cell(row=R_ppe, column=3 + i).font = BLACK
    ws3.cell(row=R_ppe, column=3 + i).number_format = CUR0
    ws3.cell(row=R_ppe, column=3 + i).border = BORDER
    # Equity roll-forward
    ws3.cell(row=R_equity_open, column=3 + i).value = f"={prev_c}{R_equity_close}"
    ws3.cell(row=R_equity_open, column=3 + i).font = BLACK
    ws3.cell(row=R_equity_open, column=3 + i).number_format = CUR0
    ws3.cell(row=R_equity_open, column=3 + i).border = BORDER
    ws3.cell(row=R_equity_close, column=3 + i).value = (
        f"={c}{R_equity_open}+'Income Statement'!{c}{R_ni}+'Cash Flow & Working Capital'!{c}{R_sbc}"
        f"+'Cash Flow & Working Capital'!{c}{R_buyback}+'Cash Flow & Working Capital'!{c}{R_div}"
    )
    ws3.cell(row=R_equity_close, column=3 + i).font = BLACK
    ws3.cell(row=R_equity_close, column=3 + i).number_format = CUR0
    ws3.cell(row=R_equity_close, column=3 + i).border = BORDER

# Balance check for every column that has a populated total-assets figure: FY26A anchor (index 2) onward
for i, c in enumerate(COLS):
    if i < 2:
        continue  # FY24A/FY25A intentionally not built out as full balance sheets (see Tab 3 note)
    ws3.cell(row=R_balcheck, column=3 + i).value = f"={c}{R_totassets}-{c}{R_totliabeq}"
    ws3.cell(row=R_balcheck, column=3 + i).font = BLACK
    ws3.cell(row=R_balcheck, column=3 + i).number_format = CUR0
    ws3.cell(row=R_balcheck, column=3 + i).border = BORDER

# ---- Wire the Income Statement's interest income and share count to the balance sheet ----
# Deliberately done here, after the Balance Sheet and Cash Flow tabs exist, so both can reference
# real rows. Both use PRIOR-period balances, which keeps the model acyclic: interest in period t
# depends only on balances closed in period t-1, so nothing feeds back into its own input.
for i, c in enumerate(COLS[HIST_N:], start=HIST_N):
    prev_c = COLS[i - 1]
    # Net interest income = prior (cash + ST investments) x yield - prior debt x blended rate
    f_int = (f"=('Balance Sheet'!{prev_c}{R_cash}+'Balance Sheet'!{prev_c}{R_sti})*{c}{R_cash_yield}"
             f"-'Balance Sheet'!{prev_c}{R_debt}*{c}{R_debt_rate}")
    ci = ws2.cell(row=R_int, column=3 + i, value=f_int)
    ci.font, ci.number_format, ci.border = BLACK, CUR0, BORDER
    # Diluted shares = prior + SBC-driven issuance - buyback-retired shares (buybacks are negative)
    f_sh = (f"={prev_c}{R_shares}"
            f"+('Cash Flow & Working Capital'!{c}{R_sbc}*{c}{R_sbc_dil_factor})/Cover_Price"
            f"+'Cash Flow & Working Capital'!{c}{R_buyback}/Cover_Price")
    cs = ws2.cell(row=R_shares, column=3 + i, value=f_sh)
    cs.font, cs.number_format, cs.border = BLACK, NUM0, BORDER

# ============================================================================
# TAB 5 — DCF & WACC
# ============================================================================
ws5 = wb.create_sheet("DCF & WACC")
set_col_widths(ws5, [2, 34] + [13] * 8 + [3])
ws5.sheet_view.showGridLines = False

# TICKER-SPECIFIC: replace every yellow WACC input with real, sourced, dated figures.
section_bar(ws5, 2, "WACC BUILD", 6)
r = 3
cell(ws5, r, 2, "Risk-free rate (10-yr US Treasury)", BLACK)
cell(ws5, r, 4, 0.0450, BLUE, PCT1, fill=YELLOW_FILL)
cell(ws5, r, 6, "[source, date]", NOTE)
RF = f"D{r}"
r += 1
cell(ws5, r, 2, "Equity risk premium (Damodaran, forward-looking)", BLACK)
cell(ws5, r, 4, 0.045, BLUE, PCT1, fill=YELLOW_FILL)
cell(ws5, r, 6, "[Damodaran latest update, date]", NOTE)
ERP = f"D{r}"
r += 1
cell(ws5, r, 2, "Beta (5-year)", BLACK)
cell(ws5, r, 4, 1.00, BLUE, "0.00", fill=YELLOW_FILL)
cell(ws5, r, 6, "[source — cross-check vs. a second data provider, methodologies differ materially]", NOTE)
BETA = f"D{r}"
r += 1
cell(ws5, r, 2, "Cost of equity (CAPM = Rf + Beta x ERP)", BOLD)
cell(ws5, r, 4, f"={RF}+{BETA}*{ERP}", BLACK, PCT1)
COE = f"D{r}"
r += 2
cell(ws5, r, 2, "Pre-tax cost of debt", BLACK)
cell(ws5, r, 4, 0.060, BLUE, PCT1, fill=YELLOW_FILL)
cell(ws5, r, 6, "[real coupon/yield on the company's actual debt, or a credit-rating-based proxy, cited]", NOTE)
KD_PRETAX = f"D{r}"
r += 1
cell(ws5, r, 2, "Marginal tax rate (for debt tax shield)", BLACK)
cell(ws5, r, 4, 0.21, BLUE, PCT1, fill=YELLOW_FILL)
cell(ws5, r, 6, "US federal statutory rate", NOTE)
TAXM = f"D{r}"
r += 1
cell(ws5, r, 2, "After-tax cost of debt", BOLD)
cell(ws5, r, 4, f"={KD_PRETAX}*(1-{TAXM})", BLACK, PCT1)
COD = f"D{r}"
r += 2
cell(ws5, r, 2, "Market value of equity ($mm)", BLACK)
cell(ws5, r, 4, "=Cover_MktCap", BLACK, CUR0)
MVE = f"D{r}"
r += 1
cell(ws5, r, 2, "Market value of debt ($mm)", BLACK)
cell(ws5, r, 4, "=Cover_Debt", BLACK, CUR0)
MVD = f"D{r}"
r += 1
cell(ws5, r, 2, "Weight of equity", BLACK)
cell(ws5, r, 4, f"={MVE}/({MVE}+{MVD})", BLACK, PCT1)
WE = f"D{r}"
r += 1
cell(ws5, r, 2, "Weight of debt", BLACK)
cell(ws5, r, 4, f"={MVD}/({MVE}+{MVD})", BLACK, PCT1)
WD = f"D{r}"
r += 1
cell(ws5, r, 2, "WACC", BOLD, fill=LIGHT_FILL)
cell(ws5, r, 4, f"={WE}*{COE}+{WD}*{COD}", BOLD, PCT1, fill=LIGHT_FILL)
WACC_CELL = f"D{r}"
r += 1
# TICKER-SPECIFIC: rewrite this note to describe the real equity/debt weight split for this
# company and what it implies (e.g. debt-light mega-cap → WACC ≈ cost of equity; a leveraged
# industrial → debt weight materially pulls WACC down) — worth calling out either way.
cell(ws5, r, 2,
     "Note: [state the real market cap vs. debt figures and what the resulting weight split "
     "implies for WACC — e.g. whether it's effectively equal to cost of equity, or meaningfully "
     "pulled down by leverage].",
     NOTE, wrap=True)
ws5.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
ws5.row_dimensions[r].height = 40

r += 3
section_bar(ws5, r, "UNLEVERED FREE CASH FLOW BUILD (FY2027E-FY2031E)", 9)
r += 1
hdr_dcf = r
cell(ws5, hdr_dcf, 2, "", BLACK)
for i, p in enumerate(PERIODS[3:]):
    cell(ws5, hdr_dcf, 3 + i, p, HDR, fill=HDR_FILL, align="center", border=True)
proj_cols = COLS[3:]  # F..J
r += 1
R_dcf_ebit = r
for i, c in enumerate(proj_cols):
    cell(ws5, r, 3 + i, f"='Income Statement'!{c}{R_ebit}", BLACK, CUR0, border=True)
cell(ws5, r, 2, "EBIT", BLACK)
r += 1
R_dcf_nopat = r
for i, c in enumerate(proj_cols):
    dcf_col = get_column_letter(3 + i)
    cell(ws5, r, 3 + i, f"={dcf_col}{R_dcf_ebit}*(1-'Income Statement'!{c}{R_tax_rate})", BLACK, CUR0, border=True)
cell(ws5, r, 2, "NOPAT (EBIT x (1 - tax rate))", BLACK)
r += 1
R_dcf_da = r
for i, c in enumerate(proj_cols):
    cell(ws5, r, 3 + i, f"='Cash Flow & Working Capital'!{c}{R_da}", BLACK, CUR0, border=True)
cell(ws5, r, 2, "+ D&A", BLACK)
r += 1
R_dcf_capex = r
for i, c in enumerate(proj_cols):
    cell(ws5, r, 3 + i, f"=-'Cash Flow & Working Capital'!{c}{R_capex}", BLACK, CUR0, border=True)
cell(ws5, r, 2, "- Capex", BLACK)
r += 1
R_dcf_dwc = r
for i, c in enumerate(proj_cols):
    cell(ws5, r, 3 + i, f"='Cash Flow & Working Capital'!{c}{R_dwc}", BLACK, CUR0, border=True)
cell(ws5, r, 2, "- Increase in net working capital", BLACK)
r += 1
R_dcf_ufcf = r
for i, c in enumerate(proj_cols):
    dcf_col = get_column_letter(3 + i)
    cell(ws5, r, 3 + i, f"={dcf_col}{R_dcf_nopat}+{dcf_col}{R_dcf_da}+{dcf_col}{R_dcf_capex}+{dcf_col}{R_dcf_dwc}", BOLD, CUR0, border=True, fill=LIGHT_FILL)
cell(ws5, r, 2, "Unlevered free cash flow", BOLD, fill=LIGHT_FILL)
r += 1
R_dcf_period = r
for i in range(len(proj_cols)):
    cell(ws5, r, 3 + i, i + 1, BLUE, "0", border=True)
cell(ws5, r, 2, "Discount period (years)", BLACK)
r += 1
R_dcf_factor = r
for i, c in enumerate(proj_cols):
    dcf_col = get_column_letter(3 + i)
    cell(ws5, r, 3 + i, f"=1/(1+{WACC_CELL})^{dcf_col}{R_dcf_period}", BLACK, "0.000", border=True)
cell(ws5, r, 2, "Discount factor", BLACK)
r += 1
R_dcf_pv = r
for i, c in enumerate(proj_cols):
    dcf_col = get_column_letter(3 + i)
    cell(ws5, r, 3 + i, f"={dcf_col}{R_dcf_ufcf}*{dcf_col}{R_dcf_factor}", BLACK, CUR0, border=True)
cell(ws5, r, 2, "PV of unlevered FCF", BLACK)

r += 2
section_bar(ws5, r, "TERMINAL VALUE & ENTERPRISE VALUE", 6)
r += 1
# TICKER-SPECIFIC: terminal growth should not exceed a reasonable long-run nominal GDP growth
# rate (~2.5-3.5% is a common ceiling) unless there's a specific, stated reason for more
cell(ws5, r, 2, "Terminal growth rate", BLACK)
cell(ws5, r, 4, 0.025, BLUE, PCT1, fill=YELLOW_FILL)
TGR = f"D{r}"
r += 1
cell(ws5, r, 2, "Terminal year unlevered FCF (FY31E x (1+g))", BLACK)
last_col = get_column_letter(3 + len(proj_cols) - 1)
cell(ws5, r, 4, f"={last_col}{R_dcf_ufcf}*(1+{TGR})", BLACK, CUR0)
TERM_FCF = f"D{r}"
r += 1
cell(ws5, r, 2, "Terminal value (Gordon growth: FCF/(WACC-g))", BLACK)
cell(ws5, r, 4, f"={TERM_FCF}/({WACC_CELL}-{TGR})", BLACK, CUR0)
TV = f"D{r}"
r += 1
cell(ws5, r, 2, "PV of terminal value", BOLD)
cell(ws5, r, 4, f"={TV}*{last_col}{R_dcf_factor}", BOLD, CUR0)
PV_TV = f"D{r}"
r += 1
cell(ws5, r, 2, "Sum of PV of explicit-period FCF", BLACK)
first_col = get_column_letter(3)
cell(ws5, r, 4, f"=SUM({first_col}{R_dcf_pv}:{last_col}{R_dcf_pv})", BLACK, CUR0)
PV_FCF = f"D{r}"
r += 1
cell(ws5, r, 2, "Enterprise value", BOLD, fill=LIGHT_FILL)
cell(ws5, r, 4, f"={PV_FCF}+{PV_TV}", BOLD, CUR0, fill=LIGHT_FILL)
EV = f"D{r}"
r += 1
cell(ws5, r, 2, "  Terminal value as % of EV", NOTE)
cell(ws5, r, 4, f"={PV_TV}/{EV}", BLACK, PCT0)
r += 1
cell(ws5, r, 2, "Plus: net cash", BLACK)
cell(ws5, r, 4, f"=Cover_Cash-Cover_Debt", BLACK, CUR0)
NETCASH = f"D{r}"
r += 1
cell(ws5, r, 2, "Implied equity value", BOLD)
cell(ws5, r, 4, f"={EV}+{NETCASH}", BOLD, CUR0)
EQV = f"D{r}"
r += 1
cell(ws5, r, 2, "Diluted shares outstanding (mm)", BLACK)
cell(ws5, r, 4, "=Cover_Shares", BLACK, NUM0)
SHR = f"D{r}"
r += 1
cell(ws5, r, 2, "Implied share price (DCF)", BOLD, fill=LIGHT_FILL)
cell(ws5, r, 4, f"=IFERROR({EQV}/{SHR},\"-\")", BOLD, CUR2, fill=LIGHT_FILL)
DCF_PRICE = f"D{r}"
r += 1
cell(ws5, r, 2, "Upside / (downside) vs. current price", BLACK)
cell(ws5, r, 4, f"=IFERROR({DCF_PRICE}/Cover_Price-1,\"-\")", BLACK, PCT0)
r += 2
cell(ws5, r, 2,
     "Note the terminal-value share of enterprise value above: for a company still growing "
     "meaningfully faster than GDP at the end of a 5-year explicit window, the terminal value "
     "inevitably carries most of the valuation — this is normal DCF mechanics, not a modeling "
     "error, but it does mean the terminal growth rate and WACC assumptions matter enormously "
     "(see Sensitivity tab).",
     NOTE, wrap=True)
ws5.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
ws5.row_dimensions[r].height = 40

# ============================================================================
# TAB 6 — SENSITIVITY
# ============================================================================
ws6 = wb.create_sheet("Sensitivity")
set_col_widths(ws6, [2, 20, 13, 13, 13, 13, 13, 13, 3])
ws6.sheet_view.showGridLines = False
section_bar(ws6, 2, "IMPLIED SHARE PRICE — WACC x TERMINAL GROWTH RATE", 8)
cell(ws6, 3, 2,
     "Rebuilds the DCF math directly in this grid (rather than a what-if Data Table, for "
     "portability across Excel/LibreOffice) so every cell is a live formula.", NOTE, wrap=True)
ws6.merge_cells(start_row=3, start_column=2, end_row=3, end_column=9)

# TICKER-SPECIFIC: center these scenario ranges on the base-case WACC/terminal-growth values from
# the DCF & WACC tab (e.g. base WACC ± ~3.5pp in 6 steps, base terminal growth ± ~1.0pp in 6 steps)
grid_hdr = 5
cell(ws6, grid_hdr, 2, "WACC ↓ / Terminal growth →", BOLD, wrap=True)
tgr_scenarios = [0.010, 0.015, 0.020, 0.025, 0.030, 0.035]
for i, g in enumerate(tgr_scenarios):
    cell(ws6, grid_hdr, 3 + i, g, BOLD_BLUE, PCT1, fill=LIGHT_FILL, align="center")

# The WACC axis is CENTRED ON THE MODEL'S OWN COMPUTED WACC rather than a fixed 7-12% ladder, so
# the grid always brackets the base case instead of possibly sitting entirely to one side of it.
# Both axes are live cell references ($B{row} for WACC, {col}$hdr for growth) — an earlier version
# baked the WACC values into the formula strings as Python literals while ALSO writing them to the
# label column, so editing a row label silently changed nothing. Keep both axes as references.
wacc_offsets = [-0.02, -0.01, 0.0, 0.01, 0.02, 0.03]
grid_start = grid_hdr + 1
for j, off in enumerate(wacc_offsets):
    rr = grid_start + j
    sign = "+" if off >= 0 else "-"
    cell(ws6, rr, 2, f"='DCF & WACC'!{WACC_CELL}{sign}{abs(off)}", BOLD_BLUE, PCT1, fill=LIGHT_FILL, align="center")
    w = f"$B{rr}"          # live reference to this row's WACC
    for i, g in enumerate(tgr_scenarios):
        col = get_column_letter(3 + i)
        gref = f"{col}${grid_hdr}"   # live reference to this column's terminal growth rate
        pv_terms = "+".join(
            f"'DCF & WACC'!{get_column_letter(3+k)}{R_dcf_ufcf}/(1+{w})^{k+1}" for k in range(len(proj_cols))
        )
        last_ufcf_col = get_column_letter(3 + len(proj_cols) - 1)
        inner = (f"({pv_terms})"
                 f"+(('DCF & WACC'!{last_ufcf_col}{R_dcf_ufcf}*(1+{gref}))/({w}-{gref}))/(1+{w})^{len(proj_cols)}"
                 f"+Cover_Cash-Cover_Debt")
        # IF guard: Gordon growth is meaningless once g >= WACC, and would otherwise render a
        # large NEGATIVE price rather than erroring — IFERROR alone does not catch that.
        f = f"=IF({gref}>={w},\"n/m\",IFERROR(({inner})/Cover_Shares,\"n/m\"))"
        cell(ws6, rr, 3 + i, f, BLACK, CUR2, border=True)

# Heat map across the grid so the shape of the sensitivity reads at a glance.
ws6.conditional_formatting.add(
    f"C{grid_start}:{get_column_letter(2+len(tgr_scenarios))}{grid_start+len(wacc_offsets)-1}",
    ColorScaleRule(start_type="min", start_color="F8696B",
                   mid_type="percentile", mid_value=50, mid_color="FFEB84",
                   end_type="max", end_color="63BE7B"))

r = grid_start + len(wacc_offsets) + 2
# TICKER-SPECIFIC: reference the actual base-case WACC/terminal-growth values here
cell(ws6, r, 2,
     "Read this grid alongside the base case on the DCF & WACC tab (WACC ~[X]%, terminal growth "
     "[Y]%). Implied share prices swing enormously across this table — that sensitivity is a "
     "central honest finding of any DCF: precision is illusory this far out, and the exercise is "
     "more useful for understanding what the market is implicitly assuming than for producing a "
     "single confident target.", NOTE, wrap=True)
ws6.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
ws6.row_dimensions[r].height = 55

# ============================================================================
# TAB 7 — SOURCES & ASSUMPTIONS
# ============================================================================
ws7 = wb.create_sheet("Sources & Assumptions")
set_col_widths(ws7, [2, 92, 3])
ws7.sheet_view.showGridLines = False
# TICKER-SPECIFIC: this entire tab is a real citation log, not filler — every material figure
# used elsewhere in the model should trace back to a dated source cited here. Replace the
# placeholder bullets below with the real sources you actually used, following the same pattern:
# general research-date/fiscal-year-convention note, then per-statement sourcing, then market-data
# sourcing, then a "KEY ASSUMPTIONS TO INTERROGATE" section that names the real, specific
# uncertainties in this model (not generic risk-list boilerplate).
section_bar(ws7, 2, "SOURCES & ASSUMPTIONS LOG", 1)
sources = [
    "GENERAL: research compiled [date]. [State the company's fiscal year convention if it "
    "differs from the calendar year, and the most recent completed quarter/filing reflected in "
    "this model, plus anything scheduled to be reported after this model's date that is NOT yet "
    "reflected.]",
    "",
    "[Y1]-[Y3] income statement: [cite 10-K/press releases, cross-checked against a second source]",
    "[Note any corporate actions affecting comparability — stock splits, discontinued segments, "
    "accounting-standard changes — and confirm historical figures are shown on a comparable basis]",
    "Most recent quarter's balance sheet and cash flow: [cite the specific 10-Q/press release]",
    "Full-year cash flow detail (D&A, SBC, working capital, buybacks, dividends): [cite source]",
    "Market data (price, market cap, beta): [cite source, as of date]",
    "[Cite any disclosed buyback authorization size/date and dividend policy this model's "
    "capital-return assumptions are based on]",
    "[Cite any recent financing events — debt issuance, credit rating — reflected in the debt "
    "and cost-of-debt assumptions]",
    "Risk-free rate: [cite source, date]",
    "Equity risk premium: [cite source, e.g. Damodaran's latest published update]",
    "[Cite any customer/revenue concentration disclosures relevant to this company, with the "
    "filing date they came from]",
    "[Cite sources for the industry/competitive/regulatory context referenced in the model notes]",
    "",
    "KEY ASSUMPTIONS TO INTERROGATE BEFORE USING THIS MODEL FOR A REAL DECISION:",
    "  - [Name the specific revenue growth assumption and the specific hard data point(s) — most "
    "recent actual quarter, next-quarter guidance — it was calibrated against, or flag if it is "
    "a less-grounded illustrative estimate.]",
    "  - Gross margin, R&D%, SG&A%, capex%, D&A, SBC, and buyback-pace projections are "
    "illustrative extrapolations from recent trends unless a specific citation above says "
    "otherwise. Every one of these is a yellow cell — treat this as a framework to stress-test "
    "your own assumptions, not a price target.",
    "  - [If applicable: name and quantify any large non-operating item excluded from the DCF "
    "build, and note that reported net income will not tie to this model's operating net income "
    "in years where that item is large.]",
    "  - Terminal value is a [state actual %] share of enterprise value (see DCF tab) — the "
    "higher this is, the more the terminal-growth-rate assumption matters relative to the entire "
    "explicit forecast. Lean on the Sensitivity tab, not the single base-case number.",
    "  - Beta varies by data source/methodology — cross-check against a second provider before "
    "relying on the WACC output.",
]
rr = 4
for s in sources:
    is_header = s.startswith("GENERAL") or s.startswith("KEY ASSUMPTIONS")
    cell(ws7, rr, 2, s, BOLD if is_header else NOTE, wrap=True)
    ws7.row_dimensions[rr].height = 30 if s else 8
    rr += 1

# TICKER-SPECIFIC: name the real, specific forward uncertainties for this company/sector in place
# of the bracketed text below — not generic boilerplate.
cell(ws7, rr + 1, 2,
     "This model is for educational/illustrative purposes — it is not investment advice and "
     "should not be the basis for any investment decision. [COMPANY]'s actual future results "
     "depend on [name the real, specific uncertainties — competitive dynamics, regulatory risk, "
     "customer concentration, cyclicality, etc.], all of which are genuinely uncertain as of the "
     "model date.",
     NOTE, wrap=True)
ws7.merge_cells(start_row=rr + 1, start_column=2, end_row=rr + 1, end_column=2)
ws7.row_dimensions[rr + 1].height = 40

# ============================================================================
# FINISHING TOUCHES — freeze panes so the row labels stay visible while scrolling
# an 8-period model sideways. Cheap to add, and the single thing that most makes a
# wide model feel built rather than generated.
# ============================================================================
for sheet_name in ("Income Statement", "Balance Sheet", "Cash Flow & Working Capital"):
    wb[sheet_name].freeze_panes = "C7"
wb["DCF & WACC"].freeze_panes = "C3"

# TICKER-SPECIFIC: update the output filename/path
OUTPUT_PATH = f"{TICKER}_3-Statement_DCF_Model.xlsx"
wb.save(OUTPUT_PATH)
print(f"Saved {OUTPUT_PATH}")
print("NEXT — both steps are blocking, do not deliver until both are clean:")
print(f"  1. recalculate to zero formula errors")
print(f"  2. python3 scripts/verify_model.py {OUTPUT_PATH} --price <current share price>")
