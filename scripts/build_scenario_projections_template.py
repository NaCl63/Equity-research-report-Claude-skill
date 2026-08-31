#!/usr/bin/env python3
"""
Scenario-driven integrated projections template (openpyxl).

WHEN TO USE THIS TEMPLATE:
  - The company has 1-2 dominant, identifiable drivers (a stablecoin issuer's reserve yield x
    circulation, a subscription business's net-revenue-retention x seat growth, a miner's
    commodity price x production) but a single sensitivity grid (build_xlsx_template.py's
    "Driver Sensitivity" tab) isn't enough because the user wants multi-year projections, a real
    scenario-based valuation, and a quantified (not just narrative) competitive-risk view.
  - It sits BETWEEN the two other templates in this skill: lighter than
    build_3statement_dcf_template.py (no full balance sheet / 3-statement tie-out — this only
    projects the driver-to-EBITDA bridge and an illustrative FCF proxy), heavier than the
    single-driver Driver Sensitivity tab (this builds explicit multi-year Bear/Base/Bull cases,
    not just a WACC/growth grid).
  - If the user explicitly wants a full GAAP-tie-out 3-statement model, use
    build_3statement_dcf_template.py instead. If one sensitivity grid genuinely captures the
    whole story, the lighter Driver Sensitivity tab in build_xlsx_template.py is enough.

HOW TO USE THIS FILE — copy it, rename it for your ticker, and replace every block marked
"TICKER-SPECIFIC" with real, sourced data and real driver mechanics. This was extracted from a
real rebuild of a Circle Internet Group (CRCL) model in response to a structured review that
asked for: an integrated multi-year projection (not just a blended multiple), a real scenario
valuation (not one sanity-check number), quantified competitive/distribution risk (not narrative
only), and an explicit share-count/dilution bridge. Keep the architecture — Assumptions tab as
the single source of truth for every Bear/Base/Bull driver, an integrated Projections tab that
formula-references it, a Valuation tab with at least two independent methods, and a Risk
Quantification tab — even as you swap in the target company's actual drivers.

SIX NON-OBVIOUS BUGS/LESSONS THIS TEMPLATE'S STRUCTURE ALREADY ENCODES — reintroduce them at your
peril if you restructure:

  1. DON'T GROW AN ALREADY-ELAPSED PERIOD'S RUN-RATE BY A FULL ADDITIONAL YEAR'S GROWTH RATE.
     If you're modeling FY2026E mid-way through calendar/fiscal 2026 with two quarters of actuals
     in hand, DO NOT anchor "FY2026E growth" to the H1'26 run-rate and then apply a full-year
     growth rate on top — that double-counts growth that has already happened and silently
     inflates every case (a real early draft of the source model had its "Bear" case showing
     HIGHER revenue than the prior full fiscal year despite an intentionally conservative
     assumption). Anchor growth-rate driven projections to the last genuinely COMPLETE prior
     period's actual (e.g. FY2025A year-end), matching how the company itself discloses YoY
     growth, and derive an "average balance for the year" as the average of beginning and ending
     balance — don't conflate a mid-year snapshot with a full-year anchor.
  2. WATCH COLUMN-LETTER SELF-REFERENCE IN LOOPS. A formula written to column D that means to
     reference column C's value in the SAME row is an easy off-by-one when you're building many
     near-identical formula strings in a loop (`f"=D{{rr}}*..."` written INTO column D references
     itself, not the intended input column). This produced silent #VALUE! errors in the source
     build's risk-quantification table — caught only by recalc.py, not by inspection. Double check
     which column index a cell() call's `col` argument actually writes to before referencing a
     sibling column by hardcoded letter in that same call.
  3. EXCEL SHEET NAMES ARE CAPPED AT 31 CHARACTERS. A longer name is silently truncated by
     openpyxl (with a warning easy to miss in verbose output), and every formula referencing the
     full name by string will then point at a sheet that doesn't exist under that name, throwing
     #REF!/#VALUE! errors that look unrelated to the real cause. Check every `wb.create_sheet(...)`
     name is ≤31 characters before wiring cross-sheet formulas to it.
  4. WHEN A COMPANY'S OWN DISCLOSED METRIC DEFINITION DIFFERS FROM YOUR OWN ANALYTICAL ONE (e.g.
     "distribution costs as % of total revenue" vs. "as % of the sub-line that costs relate to"),
     show BOTH, clearly labeled, with an explicit reconciliation note — don't silently pick one or
     assume they're interchangeable. This is the same principle as the reconciling-plug pattern in
     build_3statement_dcf_template.py, applied to margin/ratio metrics instead of balance-sheet
     totals.
  5. CALIBRATE THE EXIT MULTIPLE TO THE DCF'S OWN IMPLIED TERMINAL MULTIPLE — DON'T JUST FLAG THE
     GAP. An earlier version of this template picked Method 1's exit EV/EBITDA multiple and Method
     2's WACC/terminal-growth pair independently, then left a note saying "if they diverge, check
     consistency." In practice that produced a Base-case gap of ~$53 vs. ~$32 (a real user caught
     this and asked for a fix) that reflected two INCONSISTENT INPUTS, not two genuinely different
     views — Method 2's terminal value, divided by Year-3 EBITDA, backs out an implied "fair"
     terminal multiple; if Method 1's assumed exit multiple sits far above that, it's not adding
     information, it's just double-counting optimism already baked into the growth/margin
     assumptions. Fix: after building both methods, compute the DCF-implied terminal multiple
     (Method 2's PV-of-terminal-value, grossed back up by (1+WACC)^N, divided by Year-N EBITDA) and
     set the Base-case exit multiple close to it. Bear and Bull can and should still carry a real
     multiple spread around that anchor (reflecting genuinely different growth/quality profiles),
     just don't let the Base case quietly disagree with its own DCF. This template's Valuation tab
     includes a WACC x terminal-growth sensitivity grid for exactly this reason — it makes clear how
     much the "implied fair multiple" actually depends on the two DCF inputs before you anchor to
     it.
  6. SHOW A BLENDED / PROBABILITY-WEIGHTED TARGET, NOT JUST THE RAW METHOD-BY-METHOD GRID. Once
     Method 1 and Method 2 are reconciled (lesson #5), add a same-tab "Blended" row (simple average
     of the two methods) plus an "Upside/(downside) to blended" row and a probability-weighted
     single target (e.g. 25% Bear / 50% Base / 25% Bull — state the weights explicitly as this
     model's own judgment, not a derived probability) on both the Valuation tab and the Cover
     dashboard. This is how institutional processes typically summarize a wide scenario range into
     one decision-useful number without hiding the range itself. GOTCHA: get the upside/downside
     sign convention right — it's (target / current − 1), not (current / target − 1). The inverted
     version is a real mistake this template's author caught only by eyeballing output values (a
     Bear-case target far below the current price rendered as "+529% upside" instead of a ~-84%
     downside) — recalc.py's error count will NOT catch a sign inversion like this, since the
     formula is not itself invalid, just wrong. Always spot-check that a target below current price
     reads as negative (downside) and a target above reads as positive (upside).

FIVE MORE LESSONS (#7-#11) FOLDED IN AFTER A ROUND OF TIER-1 UPGRADES — an OPTIONAL but recommended
extension that adds a real tied-out mini balance sheet, a real UFCF-based DCF, an SBC & dilution
schedule, a reverse DCF, and a driver-based sensitivity grid, without going all the way to
build_3statement_dcf_template.py's full GAAP build. Triggered by a real structured review of a
delivered model that flagged: no integrated 3-statement tie-out, a DCF running off an EBITDA proxy
instead of real free cash flow, no reverse-DCF or driver-grid view, and no explicit basic-vs-diluted
share walk. Look for the "OPTIONAL TIER-1 ADDITION" / "TIER-1 UPGRADE" markers in this file's code
to see exactly where each lesson is implemented — build them together, they share row references.

  7. AVOID THE INTEREST-INCOME CIRCULARITY WITH A HELD-FLAT HISTORICAL ANCHOR, NOT A COMPOUNDING
     BALANCE. A 3-statement model normally wants interest income computed off the CURRENT period's
     cash balance — but that balance depends on net income (via the cash flow statement), which
     depends on interest income, a genuine circular reference. Excel handles this natively with
     iterative calculation turned on; this template's own verification toolchain (recalc.py, which
     runs headless LibreOffice) does not handle it reliably. Fix: compute interest income off the
     HISTORICAL ANCHOR cash balance held flat across every forecast period, not the dynamically
     rolling balance — a documented, honest simplification, not an accuracy improvement. Flag it as
     exactly that in a note next to the interest-income row; don't let it look like an oversight.
  8. STOCK-BASED COMPENSATION MUST HIT EQUITY HERE TOO. Same fix as
     build_3statement_dcf_template.py's own lesson #1, easy to re-break when building a *second*,
     lighter 3-statement variant from scratch: SBC is a non-cash add-back in the cash flow statement
     AND a credit to equity in the same period (`equity_close = equity_open + net income + SBC`).
     Miss the equity-side add-back and the balance check drifts by exactly the cumulative SBC amount
     — a distinctive, diagnostic symptom if you see it.
  9. A REAL-WORLD PASS-THROUGH ASSET RARELY EQUALS ITS OFFSETTING LIABILITY EXACTLY — RECLASSIFY
     THE GAP, DON'T FORCE A MISMATCH INTO THE FORECAST. Any business with a custodial/pass-through
     balance-sheet item (a stablecoin issuer's reserve assets vs. its circulation liability, a bank's
     customer deposits vs. corresponding cash, an insurer's segregated/restricted assets) will have
     REPORTED figures for the asset and liability sides that differ slightly (timing, cash sitting in
     a different bucket, etc.). Setting the forward model's pass-through asset line exactly equal to
     its liability line each period (a reasonable simplification — it's what makes the pass-through a
     pass-through, incapable of ever breaking the balance check) is fine, but only after reclassifying
     the real anchor-period gap into the "other operating assets" reconciling-plug line first. Skipping
     that reclassification bakes in a permanent, silent balance-sheet hole sized to the real-world gap
     — this was caught by hand-deriving the balance identity from the anchor figures BEFORE running
     the build script, not by recalc.py (which cannot see a wrong-but-internally-consistent anchor).
 10. THE DILUTED-SHARES "BEGINNING OF YEAR" ROW MUST REFERENCE LAST YEAR'S ENDING-DILUTED ROW, NOT A
     NEIGHBORING ROW. When basic and diluted share counts are tracked on adjacent rows (a real fix
     applied while building this upgrade), it is an easy off-by-one to point year N's "beginning
     diluted" formula at year N-1's ENDING BASIC row instead of its ENDING DILUTED row — both are
     plausible-looking neighbors and the error produces numbers that are wrong but not obviously so
     (diluted still exceeds basic, growth still looks monotonic). Caught by manual inspection before
     running, not by recalc.py. Comment the cross-year reference explicitly (e.g. "# prior year's
     ending diluted") so the intent is unambiguous to the next person editing it.
 11. A REVERSE DCF SOLVES ONE LEVER AT A TIME, NOT JOINTLY — AND SAY SO. "What does the current price
     imply?" has infinitely many answers if you let growth, margin, and the exit multiple all move
     together (underdetermined). Solve each lever independently, holding every other Base-case
     assumption fixed, and show all of them side by side against their own Base-case value — this
     is standard reverse-DCF practice, not a shortcut. If most single-lever answers land far outside
     anything the comps table or the company's own historical range supports, say explicitly that the
     price likely reflects several levers running above Base simultaneously — don't quietly pick
     whichever lever's required value looks least extreme and imply that's "the" answer.

After building: run recalc.py to zero formula errors. That alone does not prove the model is
sane — separately load with data_only=True and eyeball whether Bear < Base < Bull at every driver
and every output line, whether the Bear case's FY1 figure is actually more conservative than the
most recently disclosed actual/trailing trend (see bug #1 above — this is exactly the kind of
mistake a clean recalc will NOT catch), whether Method 1 and Method 2's Base-case outputs actually
sit close together after the lesson-#5 calibration, and whether the upside/downside row's sign is
right per lesson #6. If the Tier-1 upgrade (lessons #7-#11) was built, ALSO confirm the Balance
Sheet tab's "BALANCE CHECK" row reads exactly 0 for every case/year, and that diluted shares are
always >= basic shares on the SBC & Dilution Schedule tab.
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import ColorScaleRule

FONT_NAME = "Arial"
BLUE = Font(name=FONT_NAME, color="0000FF", size=10)
BLACK = Font(name=FONT_NAME, color="000000", size=10)
GREEN = Font(name=FONT_NAME, color="008000", size=10)
BOLD = Font(name=FONT_NAME, bold=True, size=10)
BOLD_BLUE = Font(name=FONT_NAME, bold=True, color="0000FF", size=10)
TITLE = Font(name=FONT_NAME, bold=True, size=16, color="1F3864")
SUBTITLE = Font(name=FONT_NAME, size=11, italic=True, color="595959")
HDR = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
SECTION = Font(name=FONT_NAME, bold=True, size=12, color="FFFFFF")
NOTE = Font(name=FONT_NAME, italic=True, size=9, color="595959")
CASE_BEAR = Font(name=FONT_NAME, bold=True, size=10, color="9C0006")
CASE_BASE = Font(name=FONT_NAME, bold=True, size=10, color="1F3864")
CASE_BULL = Font(name=FONT_NAME, bold=True, size=10, color="006100")

HDR_FILL = PatternFill("solid", fgColor="1F3864")
SECTION_FILL = PatternFill("solid", fgColor="2E5395")
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
LIGHT_FILL = PatternFill("solid", fgColor="D9E2F3")
BEAR_FILL = PatternFill("solid", fgColor="FCE4E4")
BASE_FILL = PatternFill("solid", fgColor="E4ECFC")
BULL_FILL = PatternFill("solid", fgColor="E4FCE7")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CUR0 = '$#,##0;($#,##0);"-"'
CUR2 = '$#,##0.00;($#,##0.00);"-"'
PCT1 = '0.0%;(0.0%);"-"'
PCT0 = '0%;(0%);"-"'
MULT = '0.00"x"'
NUM0 = '#,##0;(#,##0);"-"'

wb = Workbook()

CASES = ["Bear", "Base", "Bull"]
CASE_FONT = {"Bear": CASE_BEAR, "Base": CASE_BASE, "Bull": CASE_BULL}
CASE_FILL = {"Bear": BEAR_FILL, "Base": BASE_FILL, "Bull": BULL_FILL}
# TICKER-SPECIFIC: the explicit forecast horizon. Changing this list is the ONLY place the horizon
# is defined — every year loop below reads N_FORECAST_YEARS, and the DCF discounts and terminal
# value derive from it, so the two can never disagree.
#
# BE HONEST ABOUT WHAT 3 YEARS MEANS FOR THE DCF. With a 3-year explicit window, the Gordon-growth
# terminal value typically carries 80-90% of enterprise value: the DCF is then mostly an opinion
# about the exit multiple wearing a DCF's clothing. That is defensible for a company whose driver
# genuinely is not forecastable beyond three years, and indefensible if presented as a rigorous
# intrinsic valuation. Two rules follow, both enforced elsewhere in this file:
#   - The Valuation tab prints "Terminal value as % of EV" as a visible diagnostic. Read it. If it
#     is above ~85%, either extend YEARS or say plainly in the report that the DCF is a
#     multiple-in-disguise and let Method 1 carry the valuation argument.
#   - Extending the horizon means adding a year to this list AND adding the matching per-year
#     assumption values on the Assumptions tab (the driver, margin, opex and capex rows all carry
#     one value per forecast year). The loops adapt automatically; the assumption DATA does not.
YEARS = ["FY1E", "FY2E", "FY3E"]
N_FORECAST_YEARS = len(YEARS)


def section_bar(ws, row, text, span, start_col=1):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + span - 1)
    c = ws.cell(row=row, column=start_col, value=text)
    c.font = SECTION
    c.fill = SECTION_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def cell(ws, row, col, value, font=BLACK, fmt=None, fill=None, align=None, border=False, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if align or wrap:
        c.alignment = Alignment(horizontal=align or "general", wrap_text=wrap, vertical="center")
    if border:
        c.border = BORDER
    return c


def note_block(ws, row, text, span=6, start_col=2, height=40):
    cell(ws, row, start_col, text, NOTE, wrap=True)
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + span - 1)
    ws.row_dimensions[row].height = height


# ============================================================================
# TAB 1 — COVER / DASHBOARD
# (dashboard formulas patched in near the end, once Valuation row numbers are known —
#  see the "PATCH COVER TAB DASHBOARD" section at the bottom)
# ============================================================================
COMPANY_NAME = "COMPANY NAME, INC."
TICKER = "TICK"
EXCHANGE = "NYSE"

ws1 = wb.active
ws1.title = "Cover"
set_col_widths(ws1, [3, 30, 16, 16, 16, 16, 3])
ws1.sheet_view.showGridLines = False
cell(ws1, 2, 2, COMPANY_NAME, TITLE)
cell(ws1, 3, 2, f"{EXCHANGE}: {TICKER}  |  Scenario-Driven Equity Research Model", SUBTITLE)
cell(ws1, 4, 2, "Prepared for informational/educational purposes — not investment advice", NOTE)
cell(ws1, 5, 2, "Model date:", BOLD); cell(ws1, 5, 3, "YYYY-MM-DD", BLUE)
cell(ws1, 6, 2, "Market data as of:", BOLD); cell(ws1, 6, 3, "[date of last close used]", BLUE)

# TICKER-SPECIFIC: replace every value with real, sourced market data
section_bar(ws1, 8, "MARKET SNAPSHOT", 5)
snap = [
    ("Last price", 100.00, CUR2, "[source, date]"),
    ("52-week high / low", "$000.00 / $000.00", None, "[source]"),
    ("Basic shares outstanding (mm)", 250.0, NUM0, "[source, date]"),
    ("Weighted-avg diluted shares, most recent quarter (mm)", 260.0, NUM0,
     "[primary source, e.g. the company's own press release — this is usually HIGHER than basic "
     "shares out because it includes dilutive RSUs/options; keep both, don't collapse to one figure]"),
    ("Market capitalization ($mm)", None, CUR0, "Formula: price x basic shares out"),
    ("Cash & equivalents ($mm)", 1000, CUR0, "[source, balance sheet date]"),
    ("Total debt ($mm)", 0, CUR0, "[source — state 0 explicitly if genuinely none, don't just omit]"),
    ("Net cash ($mm)", None, CUR0, "Formula: cash - debt"),
]
price_row = 9
r = price_row
for label, val, fmt, src in snap:
    cell(ws1, r, 2, label, BLACK)
    if label.startswith("Market cap"):
        cell(ws1, r, 3, "=Cover_Price*Cover_Shares", BLACK, CUR0)
        mcap_row = r
    elif label.startswith("Net cash"):
        cell(ws1, r, 3, f"=C{cash_row}-C{debt_row}", BLACK, CUR0)
        netcash_row = r
    elif isinstance(val, str):
        cell(ws1, r, 3, val, BLUE)
    else:
        cell(ws1, r, 3, val, BLUE, fmt)
        if label.startswith("Basic shares"):
            shares_row = r
        if label.startswith("Weighted-avg diluted"):
            dilshares_row = r
        if label.startswith("Cash & equivalents"):
            cash_row = r
        if label.startswith("Total debt"):
            debt_row = r
    cell(ws1, r, 5, src, NOTE)
    r += 1

wb.defined_names["Cover_Price"] = DefinedName("Cover_Price", attr_text=f"Cover!$C${price_row}")
wb.defined_names["Cover_Shares"] = DefinedName("Cover_Shares", attr_text=f"Cover!$C${shares_row}")
wb.defined_names["Cover_DilShares"] = DefinedName("Cover_DilShares", attr_text=f"Cover!$C${dilshares_row}")
wb.defined_names["Cover_MktCap"] = DefinedName("Cover_MktCap", attr_text=f"Cover!$C${mcap_row}")
wb.defined_names["Cover_NetCash"] = DefinedName("Cover_NetCash", attr_text=f"Cover!$C${netcash_row}")

dash_row = r + 2
section_bar(ws1, dash_row, "DASHBOARD — CURRENT PRICE VS. SCENARIO TARGETS", 5)
DASH_TABLE_START = dash_row + 1  # patched with real formulas at the bottom of the script

r_after_dash = DASH_TABLE_START + 24
cell(ws1, r_after_dash, 2,
     "Tabs: Cover | Assumptions | Historical Financials | Projections | Valuation | "
     "Competitive & Risk Quant | Sources & Notes", NOTE, wrap=True)
ws1.merge_cells(start_row=r_after_dash, start_column=2, end_row=r_after_dash, end_column=6)
ws1.row_dimensions[r_after_dash].height = 28
cell(ws1, r_after_dash + 2, 2,
     "DISCLAIMER: Built from public sources as of the model date above, not a live market data "
     "terminal. Bear/Base/Bull scenario assumptions are this model's own construction unless a "
     "source is cited on the Assumptions tab — see that tab before treating any case as consensus. "
     "Nothing here is investment, legal, or tax advice.", NOTE, wrap=True)
ws1.row_dimensions[r_after_dash + 2].height = 50
ws1.merge_cells(start_row=r_after_dash + 2, start_column=2, end_row=r_after_dash + 2, end_column=6)

# ============================================================================
# TAB 2 — ASSUMPTIONS (every Bear/Base/Bull driver lives here; Projections references it)
# ============================================================================
ws2a = wb.create_sheet("Assumptions")
set_col_widths(ws2a, [2, 42, 13, 13, 13, 3])
ws2a.sheet_view.showGridLines = False
section_bar(ws2a, 2, "SCENARIO ASSUMPTIONS — CHANGE ANY YELLOW CELL TO RE-RUN THE MODEL", 5)
note_block(ws2a, 3,
     "Anchor near-term (Year 1) values to the company's own guidance where it exists — divergence "
     "across cases should show up mainly in years where guidance doesn't reach. Label anything "
     "that's this model's own illustrative construction rather than a sourced figure — don't let a "
     "guessed number look as authoritative as a disclosed one.", span=5, height=42)

hdr_row = 6
cell(ws2a, hdr_row, 2, "Driver / Case", HDR, fill=HDR_FILL, border=True)
for i, y in enumerate(YEARS):
    cell(ws2a, hdr_row, 3 + i, y, HDR, fill=HDR_FILL, align="center", border=True)
r = hdr_row + 1

ASSUMPTION_ROWS = {}  # (driver_key, case) -> row number


def assumption_block(title, driver_key, values_by_case, fmt=PCT1):
    global r
    cell(ws2a, r, 2, title, BOLD)
    r += 1
    for case in CASES:
        cell(ws2a, r, 2, f"    {case}", CASE_FONT[case])
        for i, v in enumerate(values_by_case[case]):
            cell(ws2a, r, 3 + i, v, BLUE, fmt, fill=YELLOW_FILL, border=True)
        ASSUMPTION_ROWS[(driver_key, case)] = r
        r += 1
    r += 1


# TICKER-SPECIFIC: replace this whole set with the target company's actual dominant driver(s).
# Keep the shape (a growth-rate driver feeding a volume, a spread/margin driver feeding
# profitability, an opex driver, a dilution driver) — swap in what these actually ARE for this
# business. Bear should show real stress relative to the most recently disclosed trailing trend,
# not just a smaller version of Base — sanity-check this explicitly (see bug #1 in the docstring).
assumption_block("PRIMARY VOLUME DRIVER growth (YoY, e.g. units/circulation/subscribers)", "volume_growth", {
    "Bear": [0.05, 0.04, 0.03],
    "Base": [0.15, 0.13, 0.11],
    "Bull": [0.30, 0.27, 0.24],
})
assumption_block("PRIMARY SPREAD/MARGIN DRIVER (e.g. yield, take-rate, gross margin)", "spread_margin", {
    "Bear": [0.30, 0.28, 0.26],
    "Base": [0.35, 0.35, 0.35],
    "Bull": [0.40, 0.41, 0.42],
})
assumption_block("Secondary revenue line ($mm; Year 1 fixed at guidance, growth thereafter)", "other_rev", {
    "Bear": [100, 108, 117],
    "Base": [100, 120, 144],
    "Bull": [100, 135, 182],
}, fmt=CUR0)
assumption_block("Operating expenses ($mm)", "opex", {
    "Bear": [90, 99, 109],
    "Base": [90, 95, 101],
    "Bull": [90, 93, 96],
}, fmt=CUR0)
assumption_block("Diluted share count growth (net dilution, YoY)", "dilution", {
    "Bear": [0.030, 0.030, 0.030],
    "Base": [0.020, 0.020, 0.020],
    "Bull": [0.010, 0.010, 0.010],
})

# OPTIONAL TIER-1 ADDITION (docstring lessons #7-#11): only needed if you build the Balance Sheet /
# Cash Flow Statement / SBC & Dilution Schedule tabs below. Anchor sbc_pct and da_pct to the most
# recent actual quarter's real ratios, not a guess — the same principle as every other assumption.
assumption_block("Stock-based compensation (% of Total Revenue)", "sbc_pct", {
    "Bear": [0.090, 0.090, 0.090],
    "Base": [0.070, 0.065, 0.060],
    "Bull": [0.060, 0.050, 0.040],
})
assumption_block("D&A (% of Total Revenue)", "da_pct", {
    "Bear": [0.040, 0.042, 0.043],
    "Base": [0.040, 0.043, 0.045],
    "Bull": [0.040, 0.045, 0.050],
})
assumption_block("Capex (% of Total Revenue, Tier-1 detailed — case-specific)", "capex_pct_case", {
    "Bear": [0.008, 0.008, 0.008],
    "Base": [0.009, 0.009, 0.010],
    "Bull": [0.010, 0.012, 0.014],
})

section_bar(ws2a, r, "SHARED (NON-SCENARIO) ASSUMPTIONS", 5)
r += 1
cell(ws2a, r, 2, "Illustrative effective cash tax rate", BLACK)
cell(ws2a, r, 3, 0.21, BLUE, PCT1, fill=YELLOW_FILL, border=True)
R_taxrate = r
r += 1
cell(ws2a, r, 2, "Illustrative maintenance capex (% of Total Revenue)", BLACK)
cell(ws2a, r, 3, 0.01, BLUE, PCT1, fill=YELLOW_FILL, border=True)
R_capexpct = r
r += 1
cell(ws2a, r, 2, "Corporate cash short-term yield (illustrative — Tier-1 interest income only)", BLACK)
cell(ws2a, r, 3, 0.035, BLUE, PCT1, fill=YELLOW_FILL, border=True)
cell(ws2a, r, 5, "Held flat, applied to a held-flat historical cash balance — see docstring lesson "
                  "#7 (avoids the interest-income circularity this toolchain can't recalc reliably)", NOTE)
R_corpyield = r
r += 1
cell(ws2a, r, 2, "DCF discount rate (WACC, illustrative)", BLACK)
cell(ws2a, r, 3, 0.12, BLUE, PCT1, fill=YELLOW_FILL, border=True)
cell(ws2a, r, 5, "Held constant across cases to isolate the operating-scenario effect — a fuller "
                  "build would also vary discount rate by case", NOTE)
R_wacc = r
r += 1
cell(ws2a, r, 2, "DCF terminal growth rate", BLACK)
cell(ws2a, r, 3, 0.03, BLUE, PCT1, fill=YELLOW_FILL, border=True)
R_tgr = r
r += 1
cell(ws2a, r, 2, "Exit EV/Adj. EBITDA multiple", BLACK)
for i, m in enumerate([8.0, 14.0, 20.0]):
    cell(ws2a, r, 3 + i, m, BLUE, MULT, fill=YELLOW_FILL, border=True)
R_exitmult = r
cell(ws2a, r + 1, 2, "TICKER-SPECIFIC — CALIBRATE, DON'T GUESS: after building the DCF on the "
     "Valuation tab, back out its implied terminal multiple (PV of terminal value, grossed up by "
     "(1+WACC)^N, divided by Year-N EBITDA) and set the BASE case multiple here close to it — "
     "don't pick Method 1's multiple independently of Method 2's own math (see docstring lesson "
     "#5). Bear/Bull can still carry a real spread around that anchor for genuinely different "
     "growth/quality profiles.", NOTE, wrap=True)
ws2a.row_dimensions[r + 1].height = 48
ws2a.merge_cells(start_row=r + 1, start_column=2, end_row=r + 1, end_column=5)
r += 3

# TICKER-SPECIFIC: this MUST be the last genuinely complete PRIOR fiscal period's actual, not a
# mid-current-year run-rate — see bug #1 in the docstring. If you only have a mid-year snapshot,
# use the prior full fiscal year end instead and let the Projections tab's average-balance
# convention handle the within-year approximation.
VOLUME_BASE = 100.0
cell(ws2a, r, 2, "Starting point: last complete prior fiscal year actual (primary volume driver)", BLACK)
cell(ws2a, r, 3, VOLUME_BASE, BLUE, CUR0, border=True)
cell(ws2a, r, 5, "Must be a genuine prior-FULL-YEAR figure — see docstring bug #1 before changing this", NOTE)
R_volbase = r
r += 1
cell(ws2a, r, 2, "Starting point: most recent quarter's weighted-avg diluted shares (mm)", BLACK)
cell(ws2a, r, 3, "=Cover_DilShares", BLACK, NUM0, border=True)
R_dilbase = r

print("Tab 1-2 done")

# ============================================================================
# TAB 3 — HISTORICAL FINANCIALS (brief anchor; expand per the xlsx skill's general conventions)
# ============================================================================
ws3h = wb.create_sheet("Historical Financials")
set_col_widths(ws3h, [2, 38, 13, 13, 13, 3])
ws3h.sheet_view.showGridLines = False
section_bar(ws3h, 2, "HISTORICAL FINANCIALS ($mm unless noted)", 4)
note_block(ws3h, 3, "TICKER-SPECIFIC: replace with 2-3 years of real historical actuals for the "
     "primary driver, the secondary revenue line, opex, and any company-defined margin metric "
     "worth reconciling against your own analytical version (see docstring bug #4).", span=4, height=40)
hdr_row = 5
for i, h in enumerate(["Line item", "FY-2A", "FY-1A", "Most recent qtr"], start=2):
    cell(ws3h, hdr_row, i, h, HDR, fill=HDR_FILL, align="center", border=True)
r = hdr_row + 1
rows_ = [
    ("Total revenue", [800, 950, 260]),
    ("  Secondary revenue line", [50, 80, 25]),
    ("  Primary revenue line (= Total - Secondary)", None),
]
for label, vals in rows_:
    cell(ws3h, r, 2, label, BLACK if vals is None else BLACK)
    if vals:
        for i, v in enumerate(vals):
            cell(ws3h, r, 3 + i, v, BLUE, CUR0, border=True)
    r += 1

r += 1
section_bar(ws3h, r, "BALANCE SHEET ANCHOR (most recent quarter/FY actual — OPTIONAL, TIER-1 ONLY)", 4)
r += 1
note_block(ws3h, r,
     "Only needed if you build the optional Tier-1 upgrade (Balance Sheet / Cash Flow Statement / "
     "SBC & Dilution Schedule tabs) — pull these from the same primary source as the income-statement "
     "anchors above (10-Q/10-K via SEC EDGAR, see references/sec_edgar_fred.md), not a secondary "
     "aggregator. GOTCHA if this business carries a pass-through/custodial asset backed by an "
     "offsetting liability (a stablecoin issuer's reserve assets vs. circulation liability, a bank's "
     "customer deposits vs. cash, an insurer's segregated assets): the reported asset and liability "
     "figures for that pass-through will rarely match exactly. If you force them equal in the forward "
     "model for a clean tie-out (reasonable), reclassify the real anchor-period gap into 'other "
     "operating assets' below FIRST — don't silently drop it. Verify by hand, before running the "
     "script, that Cash + Other assets − Other liabilities = Equity using the anchor figures alone "
     "(see docstring lesson #9 — this exact mismatch broke a real model's tie-out and was only caught "
     "by manual algebra before the first run).", span=4, height=120)
r += 5
cell(ws3h, r, 2, "Cash & equivalents", BLACK)
cell(ws3h, r, 3, 0.0, BLUE, CUR0, border=True)
R_bs_cash = r; r += 1
cell(ws3h, r, 2, "Other operating assets, net (reconciling plug)", BLACK)
cell(ws3h, r, 3, 0.0, BLUE, CUR0, border=True)
R_bs_otherassets = r; r += 1
cell(ws3h, r, 2, "Other liabilities (reconciling plug, held flat in the forward model)", BLACK)
cell(ws3h, r, 3, 0.0, BLUE, CUR0, border=True)
R_bs_otherliab = r; r += 1
cell(ws3h, r, 2, "Total stockholders' equity", BLACK)
cell(ws3h, r, 3, 0.0, BLUE, CUR0, border=True)
R_bs_equity = r; r += 1

print("Tab 3 done (expand with real data — this is a minimal anchor, not the full picture)")

# ============================================================================
# TAB 4 — PROJECTIONS (integrated multi-year, 3 fully-linked scenario blocks)
# ============================================================================
ws4p = wb.create_sheet("Projections")
set_col_widths(ws4p, [2, 44, 14, 14, 14, 3])
ws4p.sheet_view.showGridLines = False
section_bar(ws4p, 2, "INTEGRATED PROJECTIONS, BY SCENARIO", 5)
note_block(ws4p, 3,
     "Every yellow-cell driver here is pulled from the Assumptions tab — change a case's "
     "assumptions there and this build updates through to Adj. EBITDA, an illustrative FCF proxy, "
     "and per-share figures. 'FCF proxy' below is a simplified approximation (Adj. EBITDA less "
     "illustrative cash tax and maintenance capex) — kept as a quick sanity check. Each case block "
     "below also continues past EBITDA into D&A, EBIT, interest income, and net income (the OPTIONAL "
     "Tier-1 upgrade, docstring lessons #7-#11), which feeds the Balance Sheet / Cash Flow Statement "
     "/ SBC & Dilution Schedule tabs and a real-UFCF-based Method 2 DCF — delete those rows and the "
     "three extra tabs if the lighter FCF-proxy-only path is enough for this company's story. For a "
     "full GAAP-tie-out build instead, use build_3statement_dcf_template.py.",
     span=5, height=64)

PROJ_ROWS = {}  # (case, line_key) -> row


def proj_block(start_row, case):
    r = start_row
    cell(ws4p, r, 2, f"{case.upper()} CASE", CASE_FONT[case], fill=CASE_FILL[case])
    ws4p.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    r += 1
    for i, y in enumerate(YEARS):
        cell(ws4p, r, 3 + i, y, HDR, fill=HDR_FILL, align="center", border=True)
    r += 1

    def gref(driver_key, yi):
        row = ASSUMPTION_ROWS[(driver_key, case)]
        col = get_column_letter(3 + yi)
        return f"Assumptions!${col}${row}"

    # Primary volume driver, end-of-period, grown YoY off the true prior-year anchor
    R_eop = r
    cell(ws4p, r, 2, "Primary volume driver, end of period", BLACK, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        if yi == 0:
            f = f"=Assumptions!$C${R_volbase}*(1+{gref('volume_growth', 0)})"
        else:
            prev_col = get_column_letter(3 + yi - 1)
            f = f"={prev_col}{R_eop}*(1+{gref('volume_growth', yi)})"
        cell(ws4p, r, 3 + yi, f, BLACK, CUR0, border=True)
    PROJ_ROWS[(case, "eop")] = r; r += 1

    # Average balance for the year = avg of beginning/ending EOP — see docstring bug #1
    R_avg = r
    cell(ws4p, r, 2, "  Avg for the year (avg of beg./end. EOP)", NOTE, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        if yi == 0:
            f = f"=(Assumptions!$C${R_volbase}+{col}{R_eop})/2"
        else:
            prev_col = get_column_letter(3 + yi - 1)
            f = f"=({prev_col}{R_eop}+{col}{R_eop})/2"
        cell(ws4p, r, 3 + yi, f, BLACK, CUR0, border=True)
    PROJ_ROWS[(case, "avg")] = r; r += 1

    # Spread/margin driver
    R_marg = r
    cell(ws4p, r, 2, "Spread/margin driver", BLACK, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        cell(ws4p, r, 3 + yi, f"={gref('spread_margin', yi)}", BLACK, PCT1, border=True)
    PROJ_ROWS[(case, "margin")] = r; r += 1

    # Primary revenue = avg volume x spread/margin
    R_prim = r
    cell(ws4p, r, 2, "Primary revenue ($mm)", BLACK, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        cell(ws4p, r, 3 + yi, f"={col}{R_avg}*{col}{R_marg}", BLACK, CUR0, border=True)
    PROJ_ROWS[(case, "primrev")] = r; r += 1

    # Secondary revenue
    R_sec = r
    cell(ws4p, r, 2, "Secondary revenue ($mm)", BLACK, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        cell(ws4p, r, 3 + yi, f"={gref('other_rev', yi)}", BLACK, CUR0, border=True)
    PROJ_ROWS[(case, "secrev")] = r; r += 1

    # Total revenue
    R_tot = r
    cell(ws4p, r, 2, "Total revenue ($mm)", BOLD, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        cell(ws4p, r, 3 + yi, f"={col}{R_prim}+{col}{R_sec}", BOLD, CUR0, border=True)
    PROJ_ROWS[(case, "totrev")] = r; r += 1

    # Opex
    R_op = r
    cell(ws4p, r, 2, "Operating expenses ($mm)", BLACK, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        cell(ws4p, r, 3 + yi, f"={gref('opex', yi)}", BLACK, CUR0, border=True)
    PROJ_ROWS[(case, "opex")] = r; r += 1

    # Adj EBITDA
    R_eb = r
    cell(ws4p, r, 2, "Adjusted EBITDA ($mm)", BOLD, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        cell(ws4p, r, 3 + yi, f"={col}{R_tot}-{col}{R_op}", BOLD, CUR0, border=True)
    PROJ_ROWS[(case, "ebitda")] = r; r += 1

    cell(ws4p, r, 2, "  Adj. EBITDA margin (% of Total Revenue)", NOTE, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        cell(ws4p, r, 3 + yi, f"=IFERROR({col}{R_eb}/{col}{R_tot},\"-\")", BLACK, PCT1, border=True)
    r += 1

    # FCF proxy
    R_fcf = r
    cell(ws4p, r, 2, "Illustrative FCF proxy (EBITDA x (1-tax) - maint. capex)", BLACK, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        f = (f"={col}{R_eb}*(1-Assumptions!$C${R_taxrate})"
             f"-{col}{R_tot}*Assumptions!$C${R_capexpct}")
        cell(ws4p, r, 3 + yi, f, BLACK, CUR0, border=True)
    PROJ_ROWS[(case, "fcf")] = r; r += 1

    # Diluted shares walk-forward
    R_ds = r
    cell(ws4p, r, 2, "Diluted shares outstanding (mm)", BLACK, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        if yi == 0:
            f = f"=Assumptions!$C${R_dilbase}*(1+{gref('dilution', 0)})"
        else:
            prev_col = get_column_letter(3 + yi - 1)
            f = f"={prev_col}{R_ds}*(1+{gref('dilution', yi)})"
        cell(ws4p, r, 3 + yi, f, BLACK, NUM0, border=True)
    PROJ_ROWS[(case, "dilshares")] = r; r += 1

    cell(ws4p, r, 2, "  FCF proxy per diluted share", NOTE, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        cell(ws4p, r, 3 + yi, f"=IFERROR({col}{R_fcf}/{col}{R_ds},\"-\")", BLACK, CUR2, border=True)
    r += 2
    return r


# ============================================================================
# TAB 4B — OPTIONAL TIER-1 UPGRADE: BALANCE SHEET, CASH FLOW STATEMENT (REAL UFCF), AND
# SBC & DILUTION SCHEDULE. Docstring lessons #7-#11. Delete this whole block (and the three new
# tabs/functions below) plus the pnl_extension_block() calls in the driver loop if the lighter
# FCF-proxy-only Projections tab is enough for this company's story.
# ============================================================================
ws4bs = wb.create_sheet("Balance Sheet")
ws4cf = wb.create_sheet("Cash Flow Statement")
ws4sbc = wb.create_sheet("SBC & Dilution Schedule")
for _ws in (ws4bs, ws4cf, ws4sbc):
    set_col_widths(_ws, [2, 46, 14, 14, 14, 3])
    _ws.sheet_view.showGridLines = False

section_bar(ws4sbc, 2, "SBC & DILUTION SCHEDULE, BY SCENARIO", 5)
note_block(ws4sbc, 3,
     "Basic and diluted shares are grown at the SAME dilution-rate assumption as the Projections "
     "tab's own diluted-share walk, which holds a constant RSU/option overhang ratio — a "
     "simplification versus a full grant/vest/forfeiture waterfall, which most companies' proxy "
     "disclosures don't provide in enough granular detail to build with real numbers. SBC $ = Total "
     "Revenue x the SBC% assumption (anchor it to the most recent actual ratio, not a guess). See "
     "docstring lesson #10 for a real off-by-one bug this exact 'beginning of year' row invites.",
     span=5, height=56)

section_bar(ws4cf, 2, "CASH FLOW STATEMENT (REAL UFCF), BY SCENARIO", 5)
note_block(ws4cf, 3,
     "UFCF here = CFO - capex, built from real net income (not the EBITDA-based FCF proxy on the "
     "Projections tab). Swap this row in for Method 2's DCF on the Valuation tab (replace every "
     "PROJ_ROWS[(case,'fcf')] reference there with CF_ROWS[(case,'ufcf')]) once this Tier-1 upgrade "
     "is built — don't leave Method 2 wired to the proxy while this real figure sits unused. No debt/"
     "buyback/dividend financing activity assumed (CFF = 0) — add explicit financing lines if the "
     "company actually has them.", span=5, height=56)

section_bar(ws4bs, 2, "BALANCE SHEET (TIES OUT TO 0), BY SCENARIO", 5)
note_block(ws4bs, 3,
     "Ties to exactly 0 (Assets - Liabilities - Equity) BY CONSTRUCTION: cash rolls via the Cash "
     "Flow Statement's net change, other assets roll via prior + capex - D&A, other liabilities are "
     "held flat (see the Historical Financials anchor note), and equity rolls via prior + net income "
     "+ SBC (SBC must hit equity, not just cash — docstring lesson #8). recalc.py proves formulas "
     "evaluate, not that the statements tie — separately load with data_only=True and confirm the "
     "'BALANCE CHECK' row reads exactly 0 for every case/year before trusting this tab.",
     span=5, height=56)

SBC_ROWS = {}
CF_ROWS = {}
BS_ROWS = {}


def sbc_block(start_row, case, totrev_row):
    r = start_row
    cell(ws4sbc, r, 2, f"{case.upper()} CASE", CASE_FONT[case], fill=CASE_FILL[case])
    ws4sbc.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    r += 1
    for i, y in enumerate(YEARS):
        cell(ws4sbc, r, 3 + i, y, HDR, fill=HDR_FILL, align="center", border=True)
    r += 1

    def gref(driver_key, yi):
        row = ASSUMPTION_ROWS[(driver_key, case)]
        col = get_column_letter(3 + yi)
        return f"Assumptions!${col}${row}"

    R_dilbeg = r
    R_dilend = r + 1
    cell(ws4sbc, R_dilbeg, 2, "Diluted shares, beginning of year", BLACK, fill=CASE_FILL[case])
    cell(ws4sbc, R_dilend, 2, "Diluted shares, end of year", BLACK, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        if yi == 0:
            begf = f"=Assumptions!$C${R_dilbase}"
            endf = f"=Assumptions!$C${R_dilbase}*(1+{gref('dilution', 0)})"
        else:
            prev_col = get_column_letter(3 + yi - 1)
            begf = f"={prev_col}{R_dilend}"  # prior year's ending diluted (see docstring lesson #10)
            endf = f"={prev_col}{R_dilend}*(1+{gref('dilution', yi)})"
        cell(ws4sbc, R_dilbeg, 3 + yi, begf, BLACK, NUM0, border=True)
        cell(ws4sbc, R_dilend, 3 + yi, endf, BLACK, NUM0, border=True)
    r = R_dilend + 1

    R_sbc = r
    cell(ws4sbc, r, 2, "Stock-based compensation ($mm)", BOLD, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        cell(ws4sbc, r, 3 + yi, f"=Projections!{col}{totrev_row}*{gref('sbc_pct', yi)}", BOLD, CUR0, border=True)
    r += 2

    SBC_ROWS[(case, "dilbeg")] = R_dilbeg
    SBC_ROWS[(case, "dilshares")] = R_dilend
    SBC_ROWS[(case, "sbc")] = R_sbc
    return r


def pnl_extension_block(start_row, case, totrev_row, ebitda_row, sbc_row):
    """Continues a Projections-tab case block past EBITDA into a real net income, so Method 2's DCF
    and the Balance Sheet/Cash Flow tabs can use real UFCF instead of the EBITDA-based FCF proxy
    above. See docstring lessons #7-#9."""
    r = start_row

    def gref(driver_key, yi):
        row = ASSUMPTION_ROWS[(driver_key, case)]
        col = get_column_letter(3 + yi)
        return f"Assumptions!${col}${row}"

    R_da = r
    cell(ws4p, r, 2, "  D&A ($mm)", BLACK, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        cell(ws4p, r, 3 + yi, f"={col}{totrev_row}*{gref('da_pct', yi)}", BLACK, CUR0, border=True)
    r += 1

    R_sbcref = r
    cell(ws4p, r, 2, "  Stock-based compensation ($mm)", BLACK, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        cell(ws4p, r, 3 + yi, f"='SBC & Dilution Schedule'!{col}{sbc_row}", GREEN, CUR0, border=True)
    r += 1

    R_ebit = r
    cell(ws4p, r, 2, "EBIT ($mm)", BOLD, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        cell(ws4p, r, 3 + yi, f"={col}{ebitda_row}-{col}{R_da}-{col}{R_sbcref}", BOLD, CUR0, border=True)
    r += 1

    R_intinc = r
    cell(ws4p, r, 2, "  Interest income (held-flat historical cash x corp. yield)", BLACK, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        cell(ws4p, r, 3 + yi, f"='Historical Financials'!$C${R_bs_cash}*Assumptions!$C${R_corpyield}",
             BLACK, CUR0, border=True)
    r += 1
    cell(ws4p, r, 2, "  (held flat, not compounding — see docstring lesson #7 on the circularity fix)",
         NOTE, fill=CASE_FILL[case])
    r += 1

    R_ebt = r
    cell(ws4p, r, 2, "EBT ($mm)", BLACK, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        cell(ws4p, r, 3 + yi, f"={col}{R_ebit}+{col}{R_intinc}", BLACK, CUR0, border=True)
    r += 1

    R_tax = r
    cell(ws4p, r, 2, "  Cash taxes (= MAX(EBT,0) x tax rate)", NOTE, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        cell(ws4p, r, 3 + yi, f"=MAX({col}{R_ebt},0)*Assumptions!$C${R_taxrate}", BLACK, CUR0, border=True)
    r += 1

    R_ni = r
    cell(ws4p, r, 2, "Net income ($mm)", BOLD, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        cell(ws4p, r, 3 + yi, f"={col}{R_ebt}-{col}{R_tax}", BOLD, CUR0, border=True)
    r += 2

    PROJ_ROWS[(case, "da")] = R_da
    PROJ_ROWS[(case, "ni")] = R_ni
    return r


def cf_block(start_row, case, ni_row, da_row, sbc_row, totrev_row):
    r = start_row
    cell(ws4cf, r, 2, f"{case.upper()} CASE", CASE_FONT[case], fill=CASE_FILL[case])
    ws4cf.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    r += 1
    for i, y in enumerate(YEARS):
        cell(ws4cf, r, 3 + i, y, HDR, fill=HDR_FILL, align="center", border=True)
    r += 1

    def gref(driver_key, yi):
        row = ASSUMPTION_ROWS[(driver_key, case)]
        col = get_column_letter(3 + yi)
        return f"Assumptions!${col}${row}"

    R_ni = r
    cell(ws4cf, r, 2, "Net income ($mm)", BLACK, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        cell(ws4cf, r, 3 + yi, f"=Projections!{col}{ni_row}", GREEN, CUR0, border=True)
    r += 1

    R_da = r
    cell(ws4cf, r, 2, "  + D&A ($mm)", BLACK, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        cell(ws4cf, r, 3 + yi, f"=Projections!{col}{da_row}", GREEN, CUR0, border=True)
    r += 1

    R_sbc = r
    cell(ws4cf, r, 2, "  + Stock-based compensation ($mm)", BLACK, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        cell(ws4cf, r, 3 + yi, f"='SBC & Dilution Schedule'!{col}{sbc_row}", GREEN, CUR0, border=True)
    r += 1

    R_cfo = r
    cell(ws4cf, r, 2, "Cash flow from operations ($mm)", BOLD, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        cell(ws4cf, r, 3 + yi, f"={col}{R_ni}+{col}{R_da}+{col}{R_sbc}", BOLD, CUR0, border=True)
    r += 1

    R_capex = r
    cell(ws4cf, r, 2, "Capex ($mm)", BLACK, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        cell(ws4cf, r, 3 + yi, f"=Projections!{col}{totrev_row}*{gref('capex_pct_case', yi)}", BLACK, CUR0, border=True)
    r += 1

    R_cfi = r
    cell(ws4cf, r, 2, "Cash flow from investing ($mm)", BOLD, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        cell(ws4cf, r, 3 + yi, f"=-{col}{R_capex}", BOLD, CUR0, border=True)
    r += 1

    R_cff = r
    cell(ws4cf, r, 2, "Cash flow from financing ($mm; no debt/buyback/dividend assumed)", BOLD, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        cell(ws4cf, r, 3 + yi, 0, BOLD, CUR0, border=True)
    r += 1

    R_netchg = r
    cell(ws4cf, r, 2, "Net change in cash ($mm)", BOLD, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        cell(ws4cf, r, 3 + yi, f"={col}{R_cfo}+{col}{R_cfi}+{col}{R_cff}", BOLD, CUR0, border=True)
    r += 1

    R_begcash = r
    R_endcash = r + 1
    cell(ws4cf, R_begcash, 2, "Cash, beginning of year", BLACK, fill=CASE_FILL[case])
    cell(ws4cf, R_endcash, 2, "Cash, end of year", BOLD, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        if yi == 0:
            begf = f"='Historical Financials'!$C${R_bs_cash}"
            begfont = GREEN
        else:
            prev_col = get_column_letter(3 + yi - 1)
            begf = f"={prev_col}{R_endcash}"
            begfont = BLACK
        cell(ws4cf, R_begcash, 3 + yi, begf, begfont, CUR0, border=True)
        cell(ws4cf, R_endcash, 3 + yi, f"={col}{R_begcash}+{col}{R_netchg}", BOLD, CUR0, border=True)
    r = R_endcash + 2

    R_ufcf = r
    cell(ws4cf, r, 2, "Real unlevered FCF (CFO - capex)", BOLD, fill=LIGHT_FILL)
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        cell(ws4cf, r, 3 + yi, f"={col}{R_cfo}-{col}{R_capex}", BOLD, CUR0, fill=LIGHT_FILL, border=True)
    r += 2

    CF_ROWS[(case, "ni")] = R_ni
    CF_ROWS[(case, "da")] = R_da
    CF_ROWS[(case, "sbc")] = R_sbc
    CF_ROWS[(case, "cfo")] = R_cfo
    CF_ROWS[(case, "capex")] = R_capex
    CF_ROWS[(case, "endcash")] = R_endcash
    CF_ROWS[(case, "ufcf")] = R_ufcf
    return r


def bs_block(start_row, case, endcash_row, ni_row, sbc_row, da_row, capex_row):
    r = start_row
    cell(ws4bs, r, 2, f"{case.upper()} CASE", CASE_FONT[case], fill=CASE_FILL[case])
    ws4bs.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    r += 1
    for i, y in enumerate(YEARS):
        cell(ws4bs, r, 3 + i, y, HDR, fill=HDR_FILL, align="center", border=True)
    r += 1

    R_cash = r
    cell(ws4bs, r, 2, "Cash & equivalents ($mm)", BLACK, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        cell(ws4bs, r, 3 + yi, f"='Cash Flow Statement'!{col}{endcash_row}", GREEN, CUR0, border=True)
    r += 1

    R_oa = r
    cell(ws4bs, r, 2, "Other operating assets, net ($mm)", BLACK, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        if yi == 0:
            f = (f"='Historical Financials'!$C${R_bs_otherassets}"
                 f"+'Cash Flow Statement'!{col}{capex_row}-'Cash Flow Statement'!{col}{da_row}")
        else:
            prev_col = get_column_letter(3 + yi - 1)
            f = (f"={prev_col}{R_oa}"
                 f"+'Cash Flow Statement'!{col}{capex_row}-'Cash Flow Statement'!{col}{da_row}")
        cell(ws4bs, r, 3 + yi, f, BLACK, CUR0, border=True)
    r += 1

    R_ta = r
    cell(ws4bs, r, 2, "Total assets ($mm)", BOLD, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        cell(ws4bs, r, 3 + yi, f"={col}{R_cash}+{col}{R_oa}", BOLD, CUR0, border=True)
    r += 2

    R_ol = r
    cell(ws4bs, r, 2, "Other liabilities ($mm; held flat — see anchor note)", BLACK, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        cell(ws4bs, r, 3 + yi, f"='Historical Financials'!$C${R_bs_otherliab}", GREEN, CUR0, border=True)
    r += 1

    R_tl = r
    cell(ws4bs, r, 2, "Total liabilities ($mm; no debt assumed)", BOLD, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        cell(ws4bs, r, 3 + yi, f"={col}{R_ol}", BOLD, CUR0, border=True)
    r += 2

    R_eq = r
    cell(ws4bs, r, 2, "Stockholders' equity ($mm)", BOLD, fill=LIGHT_FILL)
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        if yi == 0:
            f = (f"='Historical Financials'!$C${R_bs_equity}"
                 f"+'Cash Flow Statement'!{col}{ni_row}+'Cash Flow Statement'!{col}{sbc_row}")
        else:
            prev_col = get_column_letter(3 + yi - 1)
            f = (f"={prev_col}{R_eq}"
                 f"+'Cash Flow Statement'!{col}{ni_row}+'Cash Flow Statement'!{col}{sbc_row}")
        cell(ws4bs, r, 3 + yi, f, BOLD, CUR0, fill=LIGHT_FILL, border=True)
    r += 2

    R_check = r
    cell(ws4bs, r, 2, "BALANCE CHECK (Assets - Liabilities - Equity, should = 0)", BOLD, fill=CASE_FILL[case])
    for yi in range(N_FORECAST_YEARS):
        col = get_column_letter(3 + yi)
        cell(ws4bs, r, 3 + yi, f"={col}{R_ta}-{col}{R_tl}-{col}{R_eq}", BOLD, NUM0, border=True)
    r += 2

    BS_ROWS[(case, "cash")] = R_cash
    BS_ROWS[(case, "totassets")] = R_ta
    BS_ROWS[(case, "totliab")] = R_tl
    BS_ROWS[(case, "equity")] = R_eq
    BS_ROWS[(case, "check")] = R_check
    return r


r = 5
r_sbc = 6
r_cf = 6
r_bs = 6
for case in CASES:
    r_after_proj = proj_block(r, case)
    r_sbc = sbc_block(r_sbc, case, totrev_row=PROJ_ROWS[(case, "totrev")])
    r = pnl_extension_block(r_after_proj, case, totrev_row=PROJ_ROWS[(case, "totrev")],
                             ebitda_row=PROJ_ROWS[(case, "ebitda")], sbc_row=SBC_ROWS[(case, "sbc")])
    r_cf = cf_block(r_cf, case, ni_row=PROJ_ROWS[(case, "ni")], da_row=PROJ_ROWS[(case, "da")],
                     sbc_row=SBC_ROWS[(case, "sbc")], totrev_row=PROJ_ROWS[(case, "totrev")])
    r_bs = bs_block(r_bs, case, endcash_row=CF_ROWS[(case, "endcash")], ni_row=CF_ROWS[(case, "ni")],
                     sbc_row=CF_ROWS[(case, "sbc")], da_row=CF_ROWS[(case, "da")],
                     capex_row=CF_ROWS[(case, "capex")])

note_block(ws4p, r + 1,
     "Sanity check before moving on: is Bear's Year-1 Total Revenue actually below the most "
     "recently disclosed trailing-period actual/run-rate? If Bear shows growth ABOVE what the "
     "company already reported, you likely have the docstring's bug #1 (growth-base double-"
     "counting) — go back to the Assumptions tab's volume-driver starting point before trusting "
     "any output below.", span=5, height=44)

print("Tab 4 (+ optional Tier-1 Balance Sheet / Cash Flow / SBC & Dilution tabs) done")

# ============================================================================
# TAB 5 — VALUATION (two independent scenario-valuation methods)
# ============================================================================
ws5v = wb.create_sheet("Valuation")
set_col_widths(ws5v, [2, 34, 14, 14, 14, 14, 3])
ws5v.sheet_view.showGridLines = False
section_bar(ws5v, 2, "SCENARIO VALUATION — METHOD 1: EXIT EV/ADJ. EBITDA MULTIPLE (YEAR 3)", 5)
r = 4
hdr3 = r
for i, c in enumerate(["", "Bear", "Base", "Bull"], start=2):
    cell(ws5v, hdr3, i, c, HDR, fill=HDR_FILL, align="center", border=True)
r += 1
R_m1_ebitda = r
cell(ws5v, r, 2, "Year-3 Adj. EBITDA ($mm)", BLACK)
for i, case in enumerate(CASES):
    cell(ws5v, r, 3 + i, f"=Projections!$E${PROJ_ROWS[(case,'ebitda')]}", GREEN, CUR0, border=True)
r += 1
R_m1_mult = r
cell(ws5v, r, 2, "Exit EV/Adj. EBITDA multiple", BLACK)
for i in range(len(CASES)):
    cell(ws5v, r, 3 + i, f"=Assumptions!${get_column_letter(3+i)}${R_exitmult}", GREEN, MULT, border=True)
r += 1
R_m1_ev = r
cell(ws5v, r, 2, "Implied enterprise value ($mm)", BLACK)
for i in range(len(CASES)):
    col = get_column_letter(3 + i)
    cell(ws5v, r, 3 + i, f"={col}{R_m1_ebitda}*{col}{R_m1_mult}", BLACK, CUR0, border=True)
r += 1
R_m1_eqv = r
cell(ws5v, r, 2, "Implied equity value ($mm, + net cash)", BOLD)
for i in range(len(CASES)):
    col = get_column_letter(3 + i)
    cell(ws5v, r, 3 + i, f"={col}{R_m1_ev}+Cover_NetCash", BOLD, CUR0, border=True)
r += 1
R_m1_shares = r
cell(ws5v, r, 2, "Year-3 diluted shares (mm)", BLACK)
for i, case in enumerate(CASES):
    cell(ws5v, r, 3 + i, f"=Projections!$E${PROJ_ROWS[(case,'dilshares')]}", GREEN, NUM0, border=True)
r += 1
R_m1_price = r
cell(ws5v, r, 2, "Implied share price (Method 1)", BOLD, fill=LIGHT_FILL)
for i in range(len(CASES)):
    col = get_column_letter(3 + i)
    cell(ws5v, r, 3 + i, f"=IFERROR({col}{R_m1_eqv}/{col}{R_m1_shares},\"-\")", BOLD, CUR2, fill=LIGHT_FILL, border=True)
r += 3

# Method 2 runs off REAL unlevered free cash flow (CFO - capex) from the Cash Flow Statement tab,
# NOT the EBITDA-based FCF proxy on the Projections tab. Those tabs are built unconditionally above,
# so this is the default and correct wiring: building a real UFCF line and then leaving the DCF
# pointed at the proxy is the worst of both worlds — the extra rigor is visible in the workbook but
# absent from the number anyone actually reads.
#
# If you strip the Tier-1 tabs out for a lighter build, set USE_REAL_UFCF = False below and the
# DCF falls back to the proxy — one switch, applied consistently to Method 2 AND the WACC x
# terminal-growth grid, so the two can never drift apart.
USE_REAL_UFCF = True


def dcf_fcf_ref(case, year_index):
    """Column reference for the cash flow feeding the DCF, for one case and one forecast year."""
    col = get_column_letter(3 + year_index)
    if USE_REAL_UFCF:
        return f"'Cash Flow Statement'!${col}${CF_ROWS[(case, 'ufcf')]}"
    return f"Projections!${col}${PROJ_ROWS[(case, 'fcf')]}"
section_bar(ws5v, r, "SCENARIO VALUATION — METHOD 2: SIMPLE MULTI-YEAR DCF", 5)
r += 1
hdr4 = r
for i, c in enumerate(["", "Bear", "Base", "Bull"], start=2):
    cell(ws5v, hdr4, i, c, HDR, fill=HDR_FILL, align="center", border=True)
r += 1
R_m2_pvfcf = r
cell(ws5v, r, 2, "PV of explicit-period unlevered FCF, discounted at WACC ($mm)"
     if USE_REAL_UFCF else "PV of explicit-period FCF proxy, discounted at WACC ($mm)", BLACK)
for i, case in enumerate(CASES):
    terms = "+".join(
        f"{dcf_fcf_ref(case, t)}/(1+Assumptions!$C${R_wacc})^{t+1}"
        for t in range(N_FORECAST_YEARS)
    )
    cell(ws5v, r, 3 + i, f"={terms}", BLACK, CUR0, border=True)
r += 1
R_m2_tv = r
cell(ws5v, r, 2, "PV of terminal value (Gordon growth on Year-3 FCF)", BLACK)
for i, case in enumerate(CASES):
    fcf_n = dcf_fcf_ref(case, N_FORECAST_YEARS - 1)
    f = (f"=({fcf_n}*(1+Assumptions!$C${R_tgr})/(Assumptions!$C${R_wacc}-Assumptions!$C${R_tgr}))"
         f"/(1+Assumptions!$C${R_wacc})^{N_FORECAST_YEARS}")
    cell(ws5v, r, 3 + i, f, BLACK, CUR0, border=True)
r += 1
R_m2_ev = r
cell(ws5v, r, 2, "Enterprise value ($mm)", BOLD)
for i in range(len(CASES)):
    col = get_column_letter(3 + i)
    cell(ws5v, r, 3 + i, f"={col}{R_m2_pvfcf}+{col}{R_m2_tv}", BOLD, CUR0, border=True)
r += 1
# DIAGNOSTIC, not decoration: with a short explicit window the terminal value carries most of the
# valuation, and a DCF that is 90% terminal value is an exit-multiple opinion in a DCF's clothing.
# Printing the share makes that visible instead of leaving it for a reviewer to work out.
R_m2_tvshare = r
cell(ws5v, r, 2, "  Terminal value as % of EV (read this — see YEARS note at top of file)", NOTE)
for i in range(len(CASES)):
    col = get_column_letter(3 + i)
    cell(ws5v, r, 3 + i, f"=IFERROR({col}{R_m2_tv}/{col}{R_m2_ev},\"n/m\")", NOTE, PCT1, border=True)
r += 1
R_m2_eqv = r
cell(ws5v, r, 2, "Implied equity value ($mm, + net cash)", BOLD)
for i in range(len(CASES)):
    col = get_column_letter(3 + i)
    cell(ws5v, r, 3 + i, f"={col}{R_m2_ev}+Cover_NetCash", BOLD, CUR0, border=True)
r += 1
R_m2_price = r
cell(ws5v, r, 2, "Implied share price (Method 2 — DCF)", BOLD, fill=LIGHT_FILL)
for i in range(len(CASES)):
    col = get_column_letter(3 + i)
    cell(ws5v, r, 3 + i, f"=IFERROR({col}{R_m2_eqv}/{col}{R_m1_shares},\"-\")", BOLD, CUR2, fill=LIGHT_FILL, border=True)
r += 3
note_block(ws5v, r,
     "CALIBRATION, not just a diagnostic note (docstring lesson #5): the Base-case exit multiple on "
     "the Assumptions tab should already have been set close to what this DCF implies as a 'fair' "
     "terminal multiple (Method 2's terminal value, grossed up by (1+WACC)^N, divided by Year-N "
     "EBITDA) — if Method 1 and Method 2's Base-case outputs still diverge by more than roughly "
     "20-30%, go back and recalibrate the Assumptions-tab multiple rather than leaving the gap and "
     "moving on. Bear/Bull can still diverge more, since those cases are meant to reflect genuinely "
     "different growth/quality assumptions, not just an uncalibrated multiple.", span=5, height=60)
r += 4

section_bar(ws5v, r, "DCF SENSITIVITY (BASE CASE) — WACC x TERMINAL GROWTH RATE", 5)
r += 1
note_block(ws5v, r,
     "Shows how much the Base-case DCF price actually swings across a reasonable range of both "
     "inputs — rebuilt directly per cell (not an Excel What-If Data Table, for LibreOffice "
     "portability). Also the tool for the calibration above: read off the price at your chosen "
     "WACC/growth pair and compare it to Method 1's output at the same multiple.", span=5, height=36)
r += 2
dcf_grid_hdr = r
cell(ws5v, dcf_grid_hdr, 2, "WACC (down) / Terminal growth (across)", BOLD, wrap=True)
tgr_scenarios = [0.020, 0.025, 0.030, 0.035, 0.040]
for i, g in enumerate(tgr_scenarios):
    cell(ws5v, dcf_grid_hdr, 3 + i, g, BLACK, PCT1, fill=LIGHT_FILL, align="center", border=True)
# WACC axis is centred on the Assumptions tab's own WACC rather than a fixed 10-14% ladder, so the
# grid always brackets the base case. Both axes are LIVE cell references — an earlier version baked
# the WACC values into the formula strings as Python literals while also writing them to the label
# column, so editing a row label silently changed nothing in the math.
wacc_offsets = [-0.02, -0.01, 0.0, 0.01, 0.02]
dcf_grid_start = dcf_grid_hdr + 1
for j, off in enumerate(wacc_offsets):
    rr2 = dcf_grid_start + j
    sign = "+" if off >= 0 else "-"
    cell(ws5v, rr2, 2, f"=Assumptions!$C${R_wacc}{sign}{abs(off)}", BLACK, PCT1,
         fill=LIGHT_FILL, align="center", border=True)
    w = f"$B{rr2}"                       # live WACC for this row
    for i, g in enumerate(tgr_scenarios):
        col = get_column_letter(3 + i)
        gref = f"{col}${dcf_grid_hdr}"   # live terminal growth for this column
        pv_terms = "+".join(
            f"{dcf_fcf_ref('Base', t)}/(1+{w})^{t+1}" for t in range(N_FORECAST_YEARS)
        )
        inner = (f"({pv_terms})"
                 f"+(({dcf_fcf_ref('Base', N_FORECAST_YEARS-1)}*(1+{gref}))"
                 f"/({w}-{gref}))/(1+{w})^{N_FORECAST_YEARS}"
                 f"+Cover_NetCash")
        # Gordon growth is meaningless once g >= WACC and would render a large NEGATIVE price
        # rather than an error, which IFERROR cannot catch. Guard it explicitly.
        f = (f"=IF({gref}>={w},\"n/m\",IFERROR(({inner})/{get_column_letter(3+1)}{R_m1_shares},\"n/m\"))")
        cell(ws5v, rr2, 3 + i, f, BLACK, CUR2, border=True)
ws5v.conditional_formatting.add(
    f"C{dcf_grid_start}:{get_column_letter(2+len(tgr_scenarios))}{dcf_grid_start+len(wacc_offsets)-1}",
    ColorScaleRule(start_type="min", start_color="F8696B",
                   mid_type="percentile", mid_value=50, mid_color="FFEB84",
                   end_type="max", end_color="63BE7B"))
r = dcf_grid_start + len(wacc_offsets) + 2

# OPTIONAL TIER-1 ADDITIONS (docstring lesson #11): a driver-based sensitivity grid and a reverse
# DCF, both anchored to the Base case's Year-3 (last-column) figures. Delete this whole block if the
# WACC x terminal-growth grid above already covers what the user needs.
EOP1_BASE = f"Projections!$C${PROJ_ROWS[('Base','eop')]}"
EOP2_BASE = f"Projections!$D${PROJ_ROWS[('Base','eop')]}"
EOP3_BASE_ACTUAL = f"Projections!$E${PROJ_ROWS[('Base','eop')]}"
AVG3_BASE = f"Projections!$E${PROJ_ROWS[('Base','avg')]}"
MARGIN3_BASE = f"Projections!$E${PROJ_ROWS[('Base','margin')]}"
SECREV3_BASE = f"Projections!$E${PROJ_ROWS[('Base','secrev')]}"
OPEX3_BASE = f"Projections!$E${PROJ_ROWS[('Base','opex')]}"
EBITDA3_BASE = f"Projections!$E${PROJ_ROWS[('Base','ebitda')]}"
SHARES3_BASE = f"Projections!$E${PROJ_ROWS[('Base','dilshares')]}"
MULT_BASE = f"Assumptions!$D${R_exitmult}"

section_bar(ws5v, r, "DRIVER SENSITIVITY (BASE CASE) — VOLUME GROWTH x MARGIN (OPTIONAL TIER-1 ADDITION)", 5)
r += 1
note_block(ws5v, r,
     "A more intuitive complement to the WACC x terminal-growth grid above, since growth and margin "
     "— not the discount rate — are usually what the bull/bear debate actually turns on. Each cell "
     "rebuilds a mini Base-case Method-1 valuation from scratch, holding a flat 2-year growth rate "
     "and the margin shown, and everything else (secondary revenue, opex, multiple, shares) at Base.",
     span=5, height=48)
r += 3


def driver_grid_price(g, margin_ref):
    eop2 = f"({EOP1_BASE}*(1+{g}))"
    eop3 = f"({eop2}*(1+{g}))"
    avg3 = f"(({eop2}+{eop3})/2)"
    primrev3 = f"({avg3}*{margin_ref})"
    totrev3 = f"({primrev3}+{SECREV3_BASE})"
    ebitda3 = f"({totrev3}-{OPEX3_BASE})"
    ev = f"({ebitda3}*{MULT_BASE})"
    eqv = f"({ev}+Cover_NetCash)"
    return f"=IFERROR({eqv}/{SHARES3_BASE},\"n/m\")"


growth_scenarios = [0.05, 0.10, 0.15, 0.20, 0.25]
margin_scenarios = [0.25, 0.30, 0.35, 0.40, 0.45]
grid_hdr = r
cell(ws5v, grid_hdr, 2, "Growth (down) / Margin (across)", BOLD, wrap=True)
for i, m in enumerate(margin_scenarios):
    cell(ws5v, grid_hdr, 3 + i, m, BLACK, PCT1, fill=LIGHT_FILL, align="center", border=True)
grid_start = grid_hdr + 1
for j, g in enumerate(growth_scenarios):
    rr = grid_start + j
    cell(ws5v, rr, 2, g, BLACK, PCT1, fill=LIGHT_FILL, align="center", border=True)
    for i, m in enumerate(margin_scenarios):
        cell(ws5v, rr, 3 + i, driver_grid_price(g, m), BLACK, CUR2, border=True)
r = grid_start + len(growth_scenarios) + 2

section_bar(ws5v, r, "REVERSE DCF — WHAT DOES THE CURRENT PRICE IMPLY? (OPTIONAL TIER-1 ADDITION)", 5)
r += 1
note_block(ws5v, r,
     "Instead of 'what is this worth', asks 'what has to be true for the current price to be right' "
     "— solves backward through Method 1, one lever at a time (volume growth CAGR, margin, or exit "
     "multiple), holding every OTHER Base-case Year-3 assumption fixed. Solving levers jointly would "
     "be underdetermined; solving one at a time is standard reverse-DCF practice and gives four "
     "honest, comparable answers instead of one unstated blend. The 'required volume CAGR' row is "
     "EXACT for this template's avg-of-beginning/ending-EOP convention (that relation is linear in "
     "end-of-Year-3 volume), not an approximation.", span=5, height=64)
r += 4

R_rev_reqebitda = r
cell(ws5v, r, 2, "Required Year-3 EBITDA to justify current price ($mm)", BLACK)
cell(ws5v, r, 3, f"=(Cover_Price*{SHARES3_BASE}-Cover_NetCash)/{MULT_BASE}", BLACK, CUR0, border=True)
cell(ws5v, r, 4, f"={EBITDA3_BASE}", NOTE, CUR0, border=True)
cell(ws5v, r, 5, "vs. this model's own Base-case Year-3 EBITDA (left of this note)", NOTE, wrap=True)
r += 1

R_rev_reqcagr = r
cell(ws5v, r, 2, "Required volume growth CAGR (holding margin/opex/multiple at Base)", BLACK)
_reqtotrev = f"(C{R_rev_reqebitda}+{OPEX3_BASE})"
_reqprimrev = f"({_reqtotrev}-{SECREV3_BASE})"
_reqavg = f"({_reqprimrev}/{MARGIN3_BASE})"
_reqeop3 = f"(2*{_reqavg}-{EOP2_BASE})"
cell(ws5v, r, 3, f"=IFERROR(({_reqeop3}/{EOP1_BASE})^(1/2)-1,\"n/m\")", BLACK, PCT1, border=True)
cell(ws5v, r, 4, f"=IFERROR(({EOP3_BASE_ACTUAL}/{EOP1_BASE})^(1/2)-1,\"n/m\")", NOTE, PCT1, border=True)
cell(ws5v, r, 5, "vs. this model's own Base-case FY1->FY3 CAGR (left of this note)", NOTE, wrap=True)
r += 1

R_rev_reqmargin = r
cell(ws5v, r, 2, "Required Year-3 spread/margin (holding volume/opex/multiple at Base)", BLACK)
cell(ws5v, r, 3, f"=IFERROR((C{R_rev_reqebitda}+{OPEX3_BASE}-{SECREV3_BASE})/{AVG3_BASE},\"n/m\")", BLACK, PCT1, border=True)
cell(ws5v, r, 4, f"={MARGIN3_BASE}", NOTE, PCT1, border=True)
cell(ws5v, r, 5, "vs. this model's own Base-case Year-3 margin (left of this note)", NOTE, wrap=True)
r += 1

R_rev_reqmult = r
cell(ws5v, r, 2, "Required exit multiple (holding volume/margin/opex at Base)", BLACK)
cell(ws5v, r, 3, f"=IFERROR((Cover_Price*{SHARES3_BASE}-Cover_NetCash)/{EBITDA3_BASE},\"n/m\")", BLACK, MULT, border=True)
cell(ws5v, r, 4, f"={MULT_BASE}", NOTE, MULT, border=True)
cell(ws5v, r, 5, "vs. this model's own Base-case exit multiple (left of this note)", NOTE, wrap=True)
r += 2
note_block(ws5v, r,
     "Read each required-value row against its Base-case counterpart in the column just to its "
     "left, not against the raw number alone — 'required multiple of 20x' only means something next "
     "to 'Base case uses 10x'. If most single-lever answers sit far beyond anything in the comps "
     "table or the company's own historical range, the price likely reflects several levers running "
     "above Base simultaneously rather than one dominant story — say so explicitly rather than "
     "picking whichever lever looks least extreme.", span=5, height=56)
r += 4

section_bar(ws5v, r, "SUMMARY VS. CURRENT PRICE & STREET TARGET", 5)
r += 1
hdr5 = r
for i, c in enumerate(["", "Bear", "Base", "Bull"], start=2):
    cell(ws5v, r, i, c, HDR, fill=HDR_FILL, align="center", border=True)
r += 1
cell(ws5v, r, 2, "Method 1 (exit multiple)", BLACK)
for i in range(len(CASES)):
    cell(ws5v, r, 3 + i, f"={get_column_letter(3+i)}{R_m1_price}", BLACK, CUR2, border=True)
R_summary_m1 = r
r += 1
cell(ws5v, r, 2, "Method 2 (DCF)", BLACK)
for i in range(len(CASES)):
    cell(ws5v, r, 3 + i, f"={get_column_letter(3+i)}{R_m2_price}", BLACK, CUR2, border=True)
R_summary_m2 = r
r += 1
R_summary_blend = r
cell(ws5v, r, 2, "Blended (avg of Method 1 & 2)", BOLD)
for i in range(len(CASES)):
    col = get_column_letter(3 + i)
    cell(ws5v, r, 3 + i, f"=AVERAGE({col}{R_summary_m1},{col}{R_summary_m2})", BOLD, CUR2, border=True)
r += 1
cell(ws5v, r, 2, "Current price", BOLD, fill=LIGHT_FILL)
for i in range(len(CASES)):
    cell(ws5v, r, 3 + i, "=Cover_Price", BOLD, CUR2, border=True, fill=LIGHT_FILL)
R_summary_current = r
r += 1
cell(ws5v, r, 2, "  Upside / (downside) to blended (blended / current - 1)", NOTE)
for i in range(len(CASES)):
    col = get_column_letter(3 + i)
    cell(ws5v, r, 3 + i, f"=IFERROR({col}{R_summary_blend}/{col}{R_summary_current}-1,\"-\")", BLACK, PCT0, border=True)
r += 1
# TICKER-SPECIFIC: replace with the real Street consensus target if available
cell(ws5v, r, 2, "Street consensus target (reference, TICKER-SPECIFIC)", NOTE)
for i in range(len(CASES)):
    cell(ws5v, r, 3 + i, 0, BLUE, CUR2, border=True, fill=YELLOW_FILL)
VALUATION_SUMMARY_START = hdr5
r += 2
R_prob_weighted = r
cell(ws5v, r, 2, "Probability-weighted target (TICKER-SPECIFIC weights — 25/50/25 shown as default)",
     BOLD, fill=LIGHT_FILL, wrap=True)
cell(ws5v, r, 3, f"=0.25*C{R_summary_blend}+0.5*D{R_summary_blend}+0.25*E{R_summary_blend}", BOLD,
     CUR2, fill=LIGHT_FILL, border=True)
cell(ws5v, r, 5, "State weights explicitly as this model's own judgment, not a derived probability "
     "— change to reflect your own view of how likely each case is.", NOTE, wrap=True)
ws5v.row_dimensions[r].height = 32

print("Tab 5 done")

# ============================================================================
# TAB 6 — COMPETITIVE & RISK QUANTIFICATION
# ============================================================================
ws6r = wb.create_sheet("Competitive & Risk Quant")  # 25 chars, well under the 31-char cap
set_col_widths(ws6r, [2, 36, 14, 14, 14, 14, 3])
ws6r.sheet_view.showGridLines = False
section_bar(ws6r, 2, "KEY RISK SENSITIVITY (TICKER-SPECIFIC: the single biggest non-driver lever)", 6)
note_block(ws6r, 3,
     "TICKER-SPECIFIC: identify the biggest risk to the primary margin/spread driver — a "
     "distribution-partner take-rate, a customer-concentration exposure, a regulatory cost — and "
     "quantify a range of outcomes here rather than leaving it as narrative only, even when no "
     "third-party estimate exists for the scenario-level impact (label it as this model's own "
     "illustrative construction when that's the case — see docstring's overall honesty principle).",
     span=6, height=48)
hdr = 6
for i, h in enumerate(["Scenario", "Base-case revenue driver ($mm)", "Post-shock value ($mm)",
                        "EBITDA impact vs. current ($mm)"], start=2):
    cell(ws6r, hdr, i, h, HDR, fill=HDR_FILL, align="center", border=True, wrap=True)
r = hdr + 1
risk_scenarios = [("Current", 0.0), ("+5pt stress", 0.05), ("+10pt stress", 0.10), ("-5pt improvement", -0.05)]
R_base_marg = PROJ_ROWS[("Base", "margin")]
R_base_revdriver = PROJ_ROWS[("Base", "primrev")]
for label, delta in risk_scenarios:
    cell(ws6r, r, 2, label, BLACK, border=True)
    cell(ws6r, r, 3, f"=Projections!$C${R_base_revdriver}", GREEN, CUR0, border=True)
    ratio_f = f"(Projections!$C${R_base_marg}+{delta})"
    cell(ws6r, r, 4, f"=Projections!$C${PROJ_ROWS[('Base','avg')]}*{ratio_f}", BLACK, CUR0, border=True)
    if delta == 0.0:
        base_row = r
        cell(ws6r, r, 5, 0, BLACK, CUR0, border=True)
    else:
        cell(ws6r, r, 5, f"=D{r}-D{base_row}", BLACK, CUR0, border=True)
    r += 1
r += 2
section_bar(ws6r, r, "MARKET/COMPETITIVE SHARE-LOSS SCENARIO (ILLUSTRATIVE)", 6)
r += 1
note_block(ws6r, r,
     "TICKER-SPECIFIC: if a specific competitive threat exists (a rival product, a large customer "
     "at risk, a new entrant), quantify a range of share/volume loss and its dollar impact here — "
     "even an illustrative, clearly-labeled range beats leaving the risk purely narrative.",
     span=6, height=36)
r += 2
shdr = r
for i, h in enumerate(["Volume lost to competitors by Year 3", "Year-3 volume impact",
                        "Year-3 revenue impact ($mm, Base spread)"], start=2):
    cell(ws6r, shdr, i, h, HDR, fill=HDR_FILL, align="center", border=True, wrap=True)
r = shdr + 1
loss_scenarios = [0.0, 0.05, 0.10, 0.20]
R_base_eop3 = PROJ_ROWS[("Base", "eop")]
R_base_marg3 = PROJ_ROWS[("Base", "margin")]
for loss in loss_scenarios:
    cell(ws6r, r, 2, f"{loss:.0%}", BLACK, border=True, align="center")
    cell(ws6r, r, 3, f"=-Projections!$E${R_base_eop3}*{loss}", BLACK, CUR0, border=True)
    cell(ws6r, r, 4, f"=C{r}*Projections!$E${R_base_marg3}", BLACK, CUR0, border=True)
    r += 1
r += 2

section_bar(ws6r, r, "IS THE BEAR CASE THE RISK SCENARIO? — EXPLICIT LINKAGE", 6)
r += 1
note_block(ws6r, r,
     "Rather than adding a fourth scenario column throughout the model (which would require "
     "rebuilding every downstream tab), calibrate the Bear case on the Assumptions/Projections/"
     "Valuation tabs so it already IS the stress scenario above — i.e. it should combine slower "
     "primary-driver growth AND a worse margin/take-rate simultaneously, consistent with the "
     "shocks quantified above. The table below makes that linkage explicit rather than leaving "
     "Bear as an unexplained set of lower numbers with no visible connection to the risk tables.",
     span=6, height=52)
r += 3
bhdr = r
for i, h in enumerate(["Driver", "Base case", "Bear case", "Bear vs. Base", "Read-through"], start=2):
    cell(ws6r, bhdr, i, h, HDR, fill=HDR_FILL, align="center", border=True, wrap=True)
ws6r.row_dimensions[bhdr].height = 28
r = bhdr + 1
# TICKER-SPECIFIC: point these at the actual driver rows on Assumptions (e.g. growth_rate,
# margin/take-rate) — the pattern below assumes ASSUMPTION_ROWS is keyed the same way as this
# template's Assumptions tab (driver_key, case) -> row.
cell(ws6r, r, 2, "Year-3 primary driver growth (YoY) [TICKER-SPECIFIC key]", BLACK, border=True)
cell(ws6r, r, 3, 0, BLUE, PCT1, fill=YELLOW_FILL, border=True)
cell(ws6r, r, 4, 0, BLUE, PCT1, fill=YELLOW_FILL, border=True)
cell(ws6r, r, 5, f"=D{r}-C{r}", BLACK, PCT1, border=True)
cell(ws6r, r, 6, "Slower growth = the volume/share-loss side of the risk quantified above, made "
     "explicit rather than left as an unexplained Bear-case number", NOTE, wrap=True)
ws6r.row_dimensions[r].height = 28
r += 1
cell(ws6r, r, 2, "Year-3 margin / take-rate [TICKER-SPECIFIC key]", BLACK, border=True)
cell(ws6r, r, 3, 0, BLUE, PCT1, fill=YELLOW_FILL, border=True)
cell(ws6r, r, 4, 0, BLUE, PCT1, fill=YELLOW_FILL, border=True)
cell(ws6r, r, 5, f"=D{r}-C{r}", BLACK, PCT1, border=True)
cell(ws6r, r, 6, "Lower margin = the take-rate/pricing side of the risk quantified above", NOTE, wrap=True)
ws6r.row_dimensions[r].height = 28
r += 2
note_block(ws6r, r,
     "If you want a materially worse outcome than the Bear case captures, the fastest way to stress "
     "it further without restructuring the model is to overwrite the Bear-case yellow cells on the "
     "Assumptions tab directly — every downstream tab recalculates automatically from those cells.",
     span=6, height=36)

print("Tab 6 done")

# ============================================================================
# TAB 7 — SOURCES & NOTES
# ============================================================================
ws7s = wb.create_sheet("Sources & Notes")
set_col_widths(ws7s, [2, 92, 3])
ws7s.sheet_view.showGridLines = False
section_bar(ws7s, 2, "SOURCES & DATA-QUALITY NOTES", 1)
sources = [
    "GENERAL: research compiled [date]. State the market-data snapshot date separately from the "
    "research-compilation date if they differ — this stock may have moved materially between them "
    "(flag explicitly on a Technical/Market tab if built, per the main SKILL.md conventions).",
    "",
    "[Cite every primary source used for historical financials, market data, and any company-"
    "defined metric reconciled on the Historical Financials tab.]",
    "",
    "KEY DATA-QUALITY FLAGS:",
    "  - [Name any metric with two valid definitions in circulation — this model's own analytical "
    "one vs. the company's disclosed one — and point to the reconciliation note, per docstring bug #4.]",
    "  - Bear/Base/Bull assumptions beyond the current-guided period are this model's own "
    "illustrative construction unless cited otherwise on the Assumptions tab.",
    "  - Risk-quantification scenarios on the Competitive & Risk Quant tab are explicitly "
    "illustrative placeholders unless a specific source is cited — replace with a real estimate "
    "the moment one is disclosed or credibly published.",
]
r = 4
for s in sources:
    is_header = s.startswith("GENERAL") or s.startswith("KEY DATA")
    cell(ws7s, r, 2, s, BOLD if is_header else NOTE, wrap=True)
    ws7s.row_dimensions[r].height = 30 if s else 8
    r += 1

# ============================================================================
# PATCH COVER TAB DASHBOARD
# ============================================================================
dr = DASH_TABLE_START
for i, c in enumerate(["", "Bear", "Base", "Bull"], start=2):
    cell(ws1, dr, i, c, HDR, fill=HDR_FILL, align="center", border=True)
dr += 1
cell(ws1, dr, 2, "Method 1 — exit multiple", BLACK)
for i in range(len(CASES)):
    col = get_column_letter(3 + i)
    cell(ws1, dr, 3 + i, f"=Valuation!{col}{VALUATION_SUMMARY_START + 1}", GREEN, CUR2, border=True)
dr += 1
cell(ws1, dr, 2, "Method 2 — DCF", BLACK)
for i in range(len(CASES)):
    col = get_column_letter(3 + i)
    cell(ws1, dr, 3 + i, f"=Valuation!{col}{VALUATION_SUMMARY_START + 2}", GREEN, CUR2, border=True)
dr += 1
cell(ws1, dr, 2, "Blended (avg of Method 1 & 2)", BOLD)
for i in range(len(CASES)):
    col = get_column_letter(3 + i)
    cell(ws1, dr, 3 + i, f"=Valuation!{col}{R_summary_blend}", GREEN, CUR2, border=True)
dr_blend = dr
dr += 1
cell(ws1, dr, 2, "Current price", BOLD, fill=LIGHT_FILL)
for i in range(len(CASES)):
    cell(ws1, dr, 3 + i, "=Cover_Price", BOLD, CUR2, border=True, fill=LIGHT_FILL)
dr_current = dr
dr += 1
cell(ws1, dr, 2, "  Upside / (downside) to blended", NOTE)
for i in range(len(CASES)):
    col = get_column_letter(3 + i)
    cell(ws1, dr, 3 + i, f"=IFERROR({col}{dr_blend}/{col}{dr_current}-1,\"-\")", BLACK, PCT0, border=True)
dr += 1
cell(ws1, dr, 2, "Probability-weighted target (TICKER-SPECIFIC weights)", BOLD, fill=LIGHT_FILL)
cell(ws1, dr, 3, f"=Valuation!$C${R_prob_weighted}", GREEN, CUR2, fill=LIGHT_FILL, border=True)
dr += 2

# TICKER-SPECIFIC: this is the single highest-value qualitative sentence on the whole dashboard —
# don't leave it as boilerplate. Compare the current price to the Bear/Base/Bull blended targets
# above and state plainly which scenario the market looks like it's underwriting, plus what would
# have to be true for the current price to be justified (which driver/margin assumption, moving
# which direction).
cell(ws1, dr, 2, "TICKER-SPECIFIC: What the market appears to be pricing in — write 2-3 sentences "
     "comparing the current price to the Bear/Base/Bull blended targets above and naming which "
     "scenario the price looks closest to, plus what has to be true (which driver/margin moving "
     "which direction) for the current price to be justified.", NOTE, wrap=True)
ws1.row_dimensions[dr].height = 58
ws1.merge_cells(start_row=dr, start_column=2, end_row=dr, end_column=6)

# TICKER-SPECIFIC: output filename. Writes to the CURRENT WORKING DIRECTORY, matching the other
# templates in this skill — an earlier version hardcoded /tmp paths (plus five intermediate debug
# saves), which scattered files outside the working directory and made the real deliverable easy
# to lose track of.
OUTPUT_PATH = f"{TICKER}_Scenario_Projections_Model.xlsx"

# Freeze the label column so scenario rows stay readable while scrolling across forecast years.
for _ws in (ws4p, ws4bs, ws4cf, ws4sbc, ws5v):
    _ws.freeze_panes = "C3"

wb.save(OUTPUT_PATH)
print(f"Saved {OUTPUT_PATH}")
print("Output is a SKELETON with placeholder data (TICKER/generic driver names).")
print("Replace every TICKER-SPECIFIC block with real data before delivering.")
print("NEXT — both steps are blocking, do not deliver until both are clean:")
print("  1. recalculate to zero formula errors")
print(f"  2. python3 scripts/verify_model.py {OUTPUT_PATH} --price <current share price>")
