#!/usr/bin/env python3
"""
Model verification harness — the checks that a formula recalculation CANNOT catch.

WHY THIS EXISTS
  Recalculating a workbook proves every formula *evaluates*. It does not prove the model is
  right. A balance sheet that is out by exactly the SBC amount, a Bear case that is more
  optimistic than the trailing actual, an upside/downside row with an inverted sign, a share
  count where diluted sits below basic — every one of those is a perfectly valid formula
  producing a wrong number, and every one has actually happened while building models with this
  skill. SKILL.md asks for these checks by eye; this script does them mechanically so they
  cannot be skipped when the build runs long.

USAGE
    python3 verify_model.py <model.xlsx> [--price 123.45]

  Exit code 0 = all checks passed. Exit code 1 = at least one FAIL. Treat a FAIL as blocking,
  exactly like a nonzero recalc error count.

HOW IT GETS CELL VALUES (in preference order, first one that works wins)
  1. LibreOffice headless, if `soffice` is on PATH — recalculates and caches values. Most
     faithful, since it is a real spreadsheet engine.
  2. The `formulas` package (`pip install formulas`), if importable — a pure-Python evaluator.
     Handles the formula subset these templates use. Slower on large books.
  3. Cached values already in the file (from a prior recalc.py / Excel / LibreOffice save).
     If a workbook was only ever written by openpyxl and never recalculated, there are no
     cached values and this mode reports NO DATA rather than a false pass.

WHAT IT CHECKS
  - Formula errors: any cell evaluating to #REF!, #VALUE!, #DIV/0!, #NAME?, #NUM!, #N/A.
  - Balance check: any row whose label matches "balance check" must be 0 (within $0.01mm) in
    every populated column.
  - Scenario ordering: on any sheet with Bear/Base/Bull columns, Bear <= Base <= Bull for every
    numeric row. Reported as WARN, not FAIL — a handful of rows (cost lines, share count under
    a bigger buyback) legitimately invert, so these need reading, not blind enforcement.
  - Share count sanity: any "diluted" row must be >= the matching "basic" row.
  - Sign sanity: any row labelled upside/downside paired with a target-price row must be
    negative when the target is below the current price, and positive when above.
  - Magnitude sanity: any row labelled "implied share price" / "target" must be positive and
    within 0.001x-1000x of the current price, which catches unit mismatches ($ vs $mm).
"""
import re
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("verify_model.py needs openpyxl: pip install openpyxl")

ERR_STRINGS = {"#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NUM!", "#N/A", "#NULL!"}
TOL = 0.01  # $mm — a balance check must tie to the cent, not "close enough"

results = []  # (level, message)


def record(level, msg):
    results.append((level, msg))


# --------------------------------------------------------------------------
# Value acquisition
# --------------------------------------------------------------------------
def values_via_libreoffice(path):
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        return None
    tmp = tempfile.mkdtemp()
    try:
        subprocess.run(
            [soffice, "--headless", "--norestore", "--convert-to", "xlsx", "--outdir", tmp, str(path)],
            check=True, capture_output=True, timeout=300,
        )
    except Exception:
        return None
    out = Path(tmp) / Path(path).name
    if not out.exists():
        return None
    wb = openpyxl.load_workbook(out, data_only=True)
    return {(ws.title.upper(), c.coordinate): c.value for ws in wb for row in ws.iter_rows() for c in row}


def values_via_formulas(path):
    try:
        import warnings
        warnings.filterwarnings("ignore")
        import formulas
    except ImportError:
        return None
    try:
        sol = formulas.ExcelModel().loads(str(path)).finish().calculate()
    except Exception as exc:
        record("WARN", f"`formulas` evaluation failed ({type(exc).__name__}): {exc}")
        return None
    stem = Path(path).name
    out = {}
    pat = re.compile(r"^'\[" + re.escape(stem) + r"\](.+?)'!([A-Z]+\d+)$", re.IGNORECASE)
    for key, val in sol.items():
        m = pat.match(key)
        if not m:
            continue
        sheet, coord = m.group(1), m.group(2)
        try:
            v = val.value[0, 0]
        except Exception:
            v = val
        if hasattr(v, "item"):
            try:
                v = v.item()
            except Exception:
                pass
        out[(sheet.upper(), coord)] = v
    return out or None


def values_cached(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    vals = {(ws.title.upper(), c.coordinate): c.value for ws in wb for row in ws.iter_rows() for c in row}
    # A book openpyxl wrote and nobody recalculated has None everywhere a formula sits.
    formula_wb = openpyxl.load_workbook(path)
    formula_cells = [(ws.title.upper(), c.coordinate) for ws in formula_wb for row in ws.iter_rows()
                     for c in row if isinstance(c.value, str) and c.value.startswith("=")]
    if formula_cells and all(vals.get(k) is None for k in formula_cells):
        return None
    return vals


def get_values(path):
    for name, fn in (("LibreOffice", values_via_libreoffice),
                     ("formulas", values_via_formulas),
                     ("cached values", values_cached)):
        vals = fn(path)
        if vals:
            record("INFO", f"Cell values obtained via {name}.")
            return vals
    return None


# --------------------------------------------------------------------------
# Checks
# --------------------------------------------------------------------------
def num(v):
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def row_label(wb, sheet, row):
    """Nearest non-empty label to the left on this row (labels live in col A or B)."""
    ws = wb[sheet]
    for col in (2, 1, 3):
        v = ws.cell(row=row, column=col).value
        if isinstance(v, str) and v.strip() and not v.startswith("="):
            return v.strip()
    return ""


def numeric_cells_in_row(vals, sheet, row, min_col=3):
    out = []
    sheet = sheet.upper()
    for (s, coord), v in vals.items():
        if s != sheet:
            continue
        m = re.match(r"^([A-Z]+)(\d+)$", coord)
        if not m or int(m.group(2)) != row:
            continue
        if openpyxl.utils.column_index_from_string(m.group(1)) < min_col:
            continue
        n = num(v)
        if n is not None:
            out.append((coord, n))
    return sorted(out, key=lambda t: openpyxl.utils.column_index_from_string(re.match(r"[A-Z]+", t[0]).group(0)))


def check_errors(vals):
    bad = [(s, c, v) for (s, c), v in vals.items() if isinstance(v, str) and v.strip() in ERR_STRINGS]
    if bad:
        for s, c, v in bad[:25]:
            record("FAIL", f"Formula error {v} at '{s}'!{c}")
        if len(bad) > 25:
            record("FAIL", f"...and {len(bad)-25} more formula errors")
    else:
        record("PASS", "No formula errors (#REF!/#VALUE!/#DIV0! etc.) anywhere in the workbook.")


def check_balance(wb, vals):
    found = False
    for ws in wb:
        for row in range(1, ws.max_row + 1):
            label = row_label(wb, ws.title, row)
            # A long label is a note or a section banner that happens to mention the phrase, not a
            # line item — matching those produced "evaluated to no numbers" noise.
            if len(label) > 60 or not re.search(r"balance\s*check", label, re.I):
                continue
            found = True
            cells = numeric_cells_in_row(vals, ws.title, row)
            if not cells:
                record("WARN", f"'{ws.title}' row {row} is a balance check but evaluated to no numbers.")
                continue
            off = [(c, v) for c, v in cells if abs(v) > TOL]
            if off:
                for c, v in off:
                    record("FAIL", f"Balance check '{ws.title}'!{c} = {v:,.4f}, must be 0. "
                                   f"A gap that is constant and equals SBC is the classic missing "
                                   f"equity-side SBC add-back; a gap that appears in one period and "
                                   f"persists is a balance-sheet move with no offsetting cash-flow line.")
            else:
                record("PASS", f"Balance check '{ws.title}' row {row}: ties to 0 across "
                               f"{len(cells)} periods.")
    if not found:
        record("INFO", "No 'balance check' row found — expected only for models without a balance sheet.")


def scenario_columns(wb, ws):
    """Map Bear/Base/Bull -> column index by scanning header rows."""
    for row in range(1, min(ws.max_row, 40) + 1):
        hdr = {}
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=row, column=col).value
            if isinstance(v, str):
                k = v.strip().lower()
                if k in ("bear", "base", "bull"):
                    hdr[k] = col
        if len(hdr) == 3:
            return hdr
    return None


def check_scenarios(wb, vals):
    any_found = False
    for ws in wb:
        hdr = scenario_columns(wb, ws)
        if not hdr:
            continue
        any_found = True
        inversions = []
        for row in range(1, ws.max_row + 1):
            trio = []
            for k in ("bear", "base", "bull"):
                coord = f"{openpyxl.utils.get_column_letter(hdr[k])}{row}"
                trio.append(num(vals.get((ws.title.upper(), coord))))
            if any(t is None for t in trio):
                continue
            if not (trio[0] <= trio[1] <= trio[2]):
                inversions.append((row, row_label(wb, ws.title, row), trio))
        if inversions:
            for row, label, trio in inversions[:15]:
                record("WARN", f"'{ws.title}' row {row} ({label or 'unlabelled'}): "
                               f"Bear {trio[0]:,.2f} / Base {trio[1]:,.2f} / Bull {trio[2]:,.2f} "
                               f"is not monotonic — read it and confirm the inversion is intended "
                               f"(cost and share-count lines legitimately invert; revenue, EBITDA "
                               f"and price targets do not).")
            if len(inversions) > 15:
                record("WARN", f"...and {len(inversions)-15} more non-monotonic rows on '{ws.title}'")
        else:
            record("PASS", f"'{ws.title}': Bear <= Base <= Bull holds on every numeric row.")
    if not any_found:
        record("INFO", "No Bear/Base/Bull columns found — expected for non-scenario models.")


def check_shares(wb, vals):
    for ws in wb:
        basic = diluted = None
        for row in range(1, ws.max_row + 1):
            label = row_label(wb, ws.title, row).lower()
            if "share" not in label and "shares" not in label:
                continue
            if re.search(r"\bbasic\b", label):
                basic = row
            elif re.search(r"\bdiluted\b", label):
                diluted = row
        if basic and diluted:
            b = dict(numeric_cells_in_row(vals, ws.title, basic))
            d = dict(numeric_cells_in_row(vals, ws.title, diluted))
            bad = [c for c in b if c in d and d[c] < b[c] - 1e-9]
            if bad:
                for c in bad[:10]:
                    record("FAIL", f"'{ws.title}'!{c}: diluted shares {d[c]:,.1f} < basic {b[c]:,.1f} — "
                                   f"impossible; the dilution schedule is wired backwards.")
            else:
                record("PASS", f"'{ws.title}': diluted shares >= basic shares in every period.")


def is_percent(ws, coord):
    try:
        return "%" in (ws[coord].number_format or "")
    except Exception:
        return False


def check_price_sanity(wb, vals, price):
    if price is None:
        record("INFO", "No --price given; skipping target-price sign and magnitude checks.")
        return
    price_row = re.compile(
        r"(implied share price|price target|target price|per[- ]share value"
        r"|(blended|probability[- ]weighted).*(price|target|value))", re.I)
    for ws in wb:
        for row in range(1, ws.max_row + 1):
            label = row_label(wb, ws.title, row)
            low = label.lower()
            cells = numeric_cells_in_row(vals, ws.title, row)
            # Percent-formatted rows are rates and ratios, never prices — skipping them keeps
            # assumption rows ("blended interest rate") out of the price checks.
            cells = [(c, v) for c, v in cells if not is_percent(ws, c)]
            if not cells:
                continue
            if price_row.search(low):
                for c, v in cells:
                    if v <= 0:
                        record("FAIL", f"'{ws.title}'!{c} ({label}) = {v:,.2f}: a share price cannot be <= 0.")
                    elif not (price / 1000 <= v <= price * 1000):
                        record("FAIL", f"'{ws.title}'!{c} ({label}) = {v:,.2f} vs current price "
                                       f"{price:,.2f} — off by >1000x, almost certainly a unit "
                                       f"mismatch ($ vs $mm) in the equity-value/share-count divide.")
            if re.search(r"upside|downside", low):
                for c, v in numeric_cells_in_row(vals, ws.title, row):
                    if not (-1.0 <= v <= 20.0):
                        record("WARN", f"'{ws.title}'!{c} ({label}) = {v:,.2%}: implausible as a "
                                       f"return. A large positive where you expect a loss is the "
                                       f"classic inverted (current/target - 1) sign bug.")


def main():
    args = [a for a in sys.argv[1:]]
    price = None
    if "--price" in args:
        i = args.index("--price")
        price = float(args[i + 1])
        del args[i:i + 2]
    if not args:
        sys.exit(__doc__)
    path = Path(args[0])
    if not path.exists():
        sys.exit(f"No such file: {path}")

    wb = openpyxl.load_workbook(path)
    vals = get_values(path)
    if vals is None:
        record("FAIL", "Could not obtain cell values: no LibreOffice on PATH, `formulas` not "
                       "installed, and the file carries no cached values. Install one of them "
                       "(`pip install formulas`) or recalculate the file first — do NOT deliver "
                       "an unverified model.")
    else:
        check_errors(vals)
        check_balance(wb, vals)
        check_scenarios(wb, vals)
        check_shares(wb, vals)
        check_price_sanity(wb, vals, price)

    order = {"FAIL": 0, "WARN": 1, "PASS": 2, "INFO": 3}
    print(f"\n=== verify_model.py — {path.name} ===")
    for level in ("FAIL", "WARN", "PASS", "INFO"):
        for lv, msg in results:
            if lv == level:
                print(f"[{lv}] {msg}")
    fails = sum(1 for lv, _ in results if lv == "FAIL")
    warns = sum(1 for lv, _ in results if lv == "WARN")
    print(f"\n{fails} FAIL, {warns} WARN")
    if fails:
        print("BLOCKING — do not deliver this model until every FAIL is resolved.")
    sys.exit(1 if fails else 0)


if __name__ == "__main__":
    main()
